from __future__ import annotations

import json
import io
import os
import shutil
import sqlite3
import threading
import time
import uuid
from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from flask import Flask, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
STORAGE_DIR = BASE_DIR / "storage"
DB_PATH = STORAGE_DIR / "system.db"

UNREGISTERED_DIR = STORAGE_DIR / "invoices" / "unregistered"
PENDING_REVIEW_DIR = STORAGE_DIR / "invoices" / "pending_review"
IN_STOCK_DIR = STORAGE_DIR / "invoices" / "in_stock"
OUT_STOCK_DIR = STORAGE_DIR / "invoices" / "out_stock"
PROCESSING_DIR = STORAGE_DIR / "invoices" / "processing"

DUPLICATE_ARCHIVE_DIR = STORAGE_DIR / "archive" / "duplicate_invoice"
PARSE_FAILED_DIR = STORAGE_DIR / "archive" / "parse_failed"
VALIDATION_FAILED_DIR = STORAGE_DIR / "archive" / "validation_failed"
REVIEW_REJECTED_DIR = STORAGE_DIR / "archive" / "review_rejected"

IN_STOCK_MASTER_DIR = STORAGE_DIR / "master" / "in_stock_master"
OUT_STOCK_MASTER_DIR = STORAGE_DIR / "master" / "out_stock_master"
REIMBURSEMENT_EXPORT_DIR = STORAGE_DIR / "master" / "reimbursement_export"
TEMP_DIR = STORAGE_DIR / "temp"

MAX_CONTENT_LENGTH = 50 * 1024 * 1024
AGENT_INTERVAL_SECONDS = int(os.getenv("AGENT_INTERVAL_SECONDS", "300"))
SECRET_KEY = os.getenv("SECRET_KEY", "material-agent-secret")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "wrprintk")
AGENT_LOCK = threading.Lock()
APP_HOST = os.getenv("APP_HOST", "0.0.0.0")
APP_PORT = int(os.getenv("APP_PORT", "5000"))

HEADER_MAP = {
    "采购日期": "purchase_date",
    "物资名称": "item_name",
    "规格型号": "spec_model",
    "单位": "unit",
    "数量": "quantity",
    "单价": "unit_price",
    "金额": "amount",
    "供应商": "supplier_name",
    "发票代码": "invoice_code",
    "发票号码": "invoice_number",
    "发票日期": "invoice_date",
    "发票金额": "invoice_amount",
    "备注": "remark",
}

REQUIRED_HEADERS = ["采购日期", "物资名称", "单位", "数量", "单价", "金额", "发票号码"]


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def build_batch_id(submitter_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"{timestamp}_{submitter_id}_{uuid.uuid4().hex[:4]}"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_number(value: Any, field_name: str, required: bool) -> float | None:
    text = normalize_text(value)
    if not text:
        if required:
            raise ValueError(f"{field_name}不能为空")
        return None
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field_name}格式错误") from exc


def normalize_date(value: Any, field_name: str, required: bool) -> str:
    if value in (None, ""):
        if required:
            raise ValueError(f"{field_name}不能为空")
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value).replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"{field_name}格式错误")


def ensure_directories() -> None:
    for path in [
        UNREGISTERED_DIR,
        PENDING_REVIEW_DIR,
        IN_STOCK_DIR,
        OUT_STOCK_DIR,
        PROCESSING_DIR,
        DUPLICATE_ARCHIVE_DIR,
        PARSE_FAILED_DIR,
        VALIDATION_FAILED_DIR,
        REVIEW_REJECTED_DIR,
        IN_STOCK_MASTER_DIR,
        OUT_STOCK_MASTER_DIR,
        REIMBURSEMENT_EXPORT_DIR,
        TEMP_DIR,
    ]:
        path.mkdir(parents=True, exist_ok=True)


@contextmanager
def db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db() -> None:
    ensure_directories()
    with db_connection() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS upload_batch (
                id TEXT PRIMARY KEY,
                team_name TEXT NOT NULL,
                submitter_name TEXT NOT NULL,
                submitter_id TEXT NOT NULL,
                submitted_at TEXT NOT NULL,
                form_file_path TEXT NOT NULL,
                remark TEXT DEFAULT '',
                status TEXT NOT NULL,
                folder_stage TEXT NOT NULL,
                review_note TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS staged_purchase_record (
                id TEXT PRIMARY KEY,
                batch_id TEXT NOT NULL,
                row_no INTEGER NOT NULL,
                purchase_date TEXT NOT NULL,
                item_name TEXT NOT NULL,
                spec_model TEXT DEFAULT '',
                unit TEXT NOT NULL,
                quantity REAL NOT NULL,
                unit_price REAL NOT NULL,
                amount REAL NOT NULL,
                supplier_name TEXT DEFAULT '',
                invoice_code TEXT DEFAULT '',
                invoice_number TEXT NOT NULL,
                invoice_date TEXT DEFAULT '',
                invoice_amount REAL,
                remark TEXT DEFAULT '',
                review_status TEXT NOT NULL,
                review_note TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS purchase_record (
                id TEXT PRIMARY KEY,
                batch_id TEXT NOT NULL,
                staged_record_id TEXT NOT NULL UNIQUE,
                row_no INTEGER NOT NULL,
                purchase_date TEXT NOT NULL,
                item_name TEXT NOT NULL,
                spec_model TEXT DEFAULT '',
                unit TEXT NOT NULL,
                quantity REAL NOT NULL,
                unit_price REAL NOT NULL,
                amount REAL NOT NULL,
                supplier_name TEXT DEFAULT '',
                invoice_code TEXT DEFAULT '',
                invoice_number TEXT NOT NULL,
                invoice_date TEXT DEFAULT '',
                invoice_amount REAL,
                remark TEXT DEFAULT '',
                stock_status TEXT NOT NULL,
                reimbursed_at TEXT DEFAULT '',
                reimbursement_batch_id TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS invoice_registry (
                id TEXT PRIMARY KEY,
                invoice_number TEXT NOT NULL UNIQUE,
                batch_id TEXT NOT NULL,
                purchase_record_id TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS reimbursement_batch (
                id TEXT PRIMARY KEY,
                batch_name TEXT NOT NULL,
                extracted_by TEXT NOT NULL,
                extracted_at TEXT NOT NULL,
                record_count INTEGER NOT NULL,
                total_amount REAL NOT NULL,
                export_file_path TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS reimbursement_record (
                id TEXT PRIMARY KEY,
                reimbursement_batch_id TEXT NOT NULL,
                purchase_record_id TEXT NOT NULL,
                invoice_number TEXT NOT NULL,
                reimbursed_at TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS process_log (
                id TEXT PRIMARY KEY,
                batch_id TEXT,
                stage TEXT NOT NULL,
                level TEXT NOT NULL,
                message TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            """
        )


def log_process(stage: str, level: str, message: str, batch_id: str | None = None) -> None:
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO process_log (id, batch_id, stage, level, message, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (uuid.uuid4().hex, batch_id, stage, level, message, now_iso()),
        )


def stage_path(stage: str) -> Path:
    mapping = {
        "unregistered": UNREGISTERED_DIR,
        "pending_review": PENDING_REVIEW_DIR,
        "in_stock": IN_STOCK_DIR,
        "out_stock": OUT_STOCK_DIR,
        "processing": PROCESSING_DIR,
        "duplicate_invoice": DUPLICATE_ARCHIVE_DIR,
        "parse_failed": PARSE_FAILED_DIR,
        "validation_failed": VALIDATION_FAILED_DIR,
        "review_rejected": REVIEW_REJECTED_DIR,
    }
    return mapping[stage]


def move_batch_folder(batch_id: str, from_stage: str, to_stage: str) -> None:
    source = stage_path(from_stage) / batch_id
    target = stage_path(to_stage) / batch_id
    if not source.exists():
        return
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists():
        shutil.rmtree(target)
    shutil.move(str(source), str(target))


def batch_form_path(batch_id: str, folder_stage: str) -> Path:
    return stage_path(folder_stage) / batch_id / "form.xlsx"


def write_meta(batch_dir: Path, meta: dict[str, Any]) -> None:
    meta_path = batch_dir / "meta.json"
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_excel(form_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(form_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("表格为空")
    headers = [normalize_text(cell) for cell in rows[0]]
    header_indexes: dict[str, int] = {}
    for required_header in REQUIRED_HEADERS:
        if required_header not in headers:
            raise ValueError(f"缺少必填列: {required_header}")
    for header_name in HEADER_MAP:
        if header_name in headers:
            header_indexes[header_name] = headers.index(header_name)

    parsed_rows: list[dict[str, Any]] = []
    for row_no, row in enumerate(rows[1:], start=2):
        row_values = list(row)
        if not any(value not in (None, "") for value in row_values):
            continue

        def cell(header_name: str) -> Any:
            index = header_indexes.get(header_name)
            if index is None or index >= len(row_values):
                return None
            return row_values[index]

        record = {
            "row_no": row_no,
            "purchase_date": normalize_date(cell("采购日期"), "采购日期", True),
            "item_name": normalize_text(cell("物资名称")),
            "spec_model": normalize_text(cell("规格型号")),
            "unit": normalize_text(cell("单位")),
            "quantity": normalize_number(cell("数量"), "数量", True),
            "unit_price": normalize_number(cell("单价"), "单价", True),
            "amount": normalize_number(cell("金额"), "金额", True),
            "supplier_name": normalize_text(cell("供应商")),
            "invoice_code": normalize_text(cell("发票代码")),
            "invoice_number": normalize_text(cell("发票号码")),
            "invoice_date": normalize_date(cell("发票日期"), "发票日期", False),
            "invoice_amount": normalize_number(cell("发票金额"), "发票金额", False),
            "remark": normalize_text(cell("备注")),
        }
        if not record["item_name"]:
            raise ValueError(f"第 {row_no} 行物资名称不能为空")
        if not record["unit"]:
            raise ValueError(f"第 {row_no} 行单位不能为空")
        if not record["invoice_number"]:
            raise ValueError(f"第 {row_no} 行发票号码不能为空")
        parsed_rows.append(record)

    if not parsed_rows:
        raise ValueError("表格没有有效数据")
    return parsed_rows


def invoice_exists(conn: sqlite3.Connection, invoice_number: str) -> tuple[bool, str]:
    pending = conn.execute(
        """
        SELECT batch_id FROM staged_purchase_record
        WHERE invoice_number = ? AND review_status = 'pending_review'
        LIMIT 1
        """,
        (invoice_number,),
    ).fetchone()
    if pending:
        return True, f"待入库阶段已存在相同发票号码，批次 {pending['batch_id']}"

    registry = conn.execute(
        "SELECT batch_id FROM invoice_registry WHERE invoice_number = ? LIMIT 1",
        (invoice_number,),
    ).fetchone()
    if registry:
        return True, f"库内已存在相同发票号码，批次 {registry['batch_id']}"

    reimbursement = conn.execute(
        """
        SELECT reimbursement_batch_id FROM reimbursement_record
        WHERE invoice_number = ? LIMIT 1
        """,
        (invoice_number,),
    ).fetchone()
    if reimbursement:
        return True, f"已存在出库报销记录，出库批次 {reimbursement['reimbursement_batch_id']}"

    return False, ""


def fail_batch(batch_id: str, from_stage: str, target_stage: str, status: str, reason: str) -> None:
    move_batch_folder(batch_id, from_stage, target_stage)
    with db_connection() as conn:
        conn.execute(
            """
            UPDATE upload_batch
            SET status = ?, folder_stage = ?, form_file_path = ?, review_note = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, target_stage, str(batch_form_path(batch_id, target_stage)), reason, now_iso(), batch_id),
        )
    log_process("agent", "error", reason, batch_id)


def process_batch(batch_row: sqlite3.Row) -> None:
    batch_id = batch_row["id"]
    move_batch_folder(batch_id, "unregistered", "processing")
    with db_connection() as conn:
        conn.execute(
            """
            UPDATE upload_batch
            SET status = 'processing', folder_stage = 'processing', form_file_path = ?, updated_at = ?
            WHERE id = ?
            """,
            (str(batch_form_path(batch_id, "processing")), now_iso(), batch_id),
        )
    form_path = PROCESSING_DIR / batch_id / "form.xlsx"
    try:
        parsed_rows = parse_excel(form_path)
    except ValueError as exc:
        fail_batch(batch_id, "processing", "validation_failed", "validation_failed", str(exc))
        return
    except Exception as exc:  # noqa: BLE001
        fail_batch(batch_id, "processing", "parse_failed", "parse_failed", f"解析失败: {exc}")
        return

    invoice_numbers = [row["invoice_number"] for row in parsed_rows]
    if len(invoice_numbers) != len(set(invoice_numbers)):
        fail_batch(batch_id, "processing", "duplicate_invoice", "duplicate_invoice", "当前表格存在重复发票号码")
        return

    with db_connection() as conn:
        for row in parsed_rows:
            exists, reason = invoice_exists(conn, row["invoice_number"])
            if exists:
                fail_batch(batch_id, "processing", "duplicate_invoice", "duplicate_invoice", reason)
                return

        current_time = now_iso()
        for row in parsed_rows:
            conn.execute(
                """
                INSERT INTO staged_purchase_record (
                    id, batch_id, row_no, purchase_date, item_name, spec_model, unit,
                    quantity, unit_price, amount, supplier_name, invoice_code,
                    invoice_number, invoice_date, invoice_amount, remark,
                    review_status, review_note, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    uuid.uuid4().hex,
                    batch_id,
                    row["row_no"],
                    row["purchase_date"],
                    row["item_name"],
                    row["spec_model"],
                    row["unit"],
                    row["quantity"],
                    row["unit_price"],
                    row["amount"],
                    row["supplier_name"],
                    row["invoice_code"],
                    row["invoice_number"],
                    row["invoice_date"],
                    row["invoice_amount"],
                    row["remark"],
                    "pending_review",
                    "",
                    current_time,
                    current_time,
                ),
            )
        conn.execute(
            """
            UPDATE upload_batch
            SET status = 'pending_review', folder_stage = 'pending_review', form_file_path = ?, review_note = '', updated_at = ?
            WHERE id = ?
            """,
            (str(batch_form_path(batch_id, "pending_review")), current_time, batch_id),
        )

    move_batch_folder(batch_id, "processing", "pending_review")
    log_process("agent", "info", "批次审核通过，已进入待入库队列", batch_id)


def run_agent_once() -> int:
    if not AGENT_LOCK.acquire(blocking=False):
        return 0
    try:
        with db_connection() as conn:
            batches = conn.execute(
                """
                SELECT * FROM upload_batch
                WHERE status = 'unregistered' AND folder_stage = 'unregistered'
                ORDER BY submitted_at ASC
                """
            ).fetchall()
        processed_count = 0
        for batch in batches:
            process_batch(batch)
            processed_count += 1
        return processed_count
    finally:
        AGENT_LOCK.release()


def confirm_staged_rows(batch_id: str, row_ids: list[str], operator: str) -> int:
    if not row_ids:
        return 0
    with db_connection() as conn:
        placeholders = ",".join("?" for _ in row_ids)
        rows = conn.execute(
            f"""
            SELECT * FROM staged_purchase_record
            WHERE batch_id = ? AND id IN ({placeholders}) AND review_status = 'pending_review'
            """,
            [batch_id, *row_ids],
        ).fetchall()
        confirmed_count = 0
        current_time = now_iso()
        for row in rows:
            purchase_record_id = uuid.uuid4().hex
            conn.execute(
                """
                INSERT INTO purchase_record (
                    id, batch_id, staged_record_id, row_no, purchase_date, item_name, spec_model,
                    unit, quantity, unit_price, amount, supplier_name, invoice_code,
                    invoice_number, invoice_date, invoice_amount, remark,
                    stock_status, reimbursed_at, reimbursement_batch_id, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    purchase_record_id,
                    row["batch_id"],
                    row["id"],
                    row["row_no"],
                    row["purchase_date"],
                    row["item_name"],
                    row["spec_model"],
                    row["unit"],
                    row["quantity"],
                    row["unit_price"],
                    row["amount"],
                    row["supplier_name"],
                    row["invoice_code"],
                    row["invoice_number"],
                    row["invoice_date"],
                    row["invoice_amount"],
                    row["remark"],
                    "in_stock",
                    "",
                    "",
                    current_time,
                    current_time,
                ),
            )
            conn.execute(
                """
                INSERT INTO invoice_registry (id, invoice_number, batch_id, purchase_record_id, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (uuid.uuid4().hex, row["invoice_number"], row["batch_id"], purchase_record_id, current_time),
            )
            conn.execute(
                """
                UPDATE staged_purchase_record
                SET review_status = 'confirmed', review_note = ?, updated_at = ?
                WHERE id = ?
                """,
                (f"已由 {operator} 确认入库", current_time, row["id"]),
            )
            confirmed_count += 1
    log_process("review", "info", f"确认入库 {confirmed_count} 行", batch_id)
    export_master_sheets()
    return confirmed_count


def reject_staged_rows(batch_id: str, row_ids: list[str], operator: str, note: str) -> int:
    if not row_ids:
        return 0
    with db_connection() as conn:
        placeholders = ",".join("?" for _ in row_ids)
        current_time = now_iso()
        result = conn.execute(
            f"""
            UPDATE staged_purchase_record
            SET review_status = 'rejected', review_note = ?, updated_at = ?
            WHERE batch_id = ? AND id IN ({placeholders}) AND review_status = 'pending_review'
            """,
            [f"{operator}打回: {note or '人工复核未通过'}", current_time, batch_id, *row_ids],
        )
        rejected_count = result.rowcount
    log_process("review", "warning", f"打回 {rejected_count} 行", batch_id)
    return rejected_count


def finalize_batch_review(batch_id: str) -> tuple[bool, str]:
    action = ""
    with db_connection() as conn:
        pending_count = conn.execute(
            "SELECT COUNT(*) AS total FROM staged_purchase_record WHERE batch_id = ? AND review_status = 'pending_review'",
            (batch_id,),
        ).fetchone()["total"]
        if pending_count:
            return False, "仍有待处理明细，不能完成复核"

        confirmed_count = conn.execute(
            "SELECT COUNT(*) AS total FROM staged_purchase_record WHERE batch_id = ? AND review_status = 'confirmed'",
            (batch_id,),
        ).fetchone()["total"]

        if confirmed_count:
            move_batch_folder(batch_id, "pending_review", "in_stock")
            conn.execute(
                """
                UPDATE upload_batch
                SET status = 'in_stock', folder_stage = 'in_stock', form_file_path = ?, updated_at = ?
                WHERE id = ?
                """,
                (str(batch_form_path(batch_id, "in_stock")), now_iso(), batch_id),
            )
            action = "in_stock"
        else:
            move_batch_folder(batch_id, "pending_review", "review_rejected")
            conn.execute(
                """
                UPDATE upload_batch
                SET status = 'review_rejected', folder_stage = 'review_rejected', form_file_path = ?, updated_at = ?
                WHERE id = ?
                """,
                (str(batch_form_path(batch_id, "review_rejected")), now_iso(), batch_id),
            )
            action = "review_rejected"

    if action == "in_stock":
        log_process("review", "info", "批次复核完成，已进入库内", batch_id)
        export_master_sheets()
        return True, "批次已进入库内"

    log_process("review", "warning", "批次全部打回，已进入异常目录", batch_id)
    return True, "批次已全部打回"


def export_master_sheets() -> None:
    with db_connection() as conn:
        in_stock_rows = conn.execute(
            """
            SELECT pr.* FROM purchase_record pr
            JOIN upload_batch ub ON ub.id = pr.batch_id
            WHERE pr.stock_status = 'in_stock' AND ub.folder_stage = 'in_stock'
            ORDER BY created_at ASC
            """
        ).fetchall()
        out_stock_rows = conn.execute(
            """
            SELECT pr.* FROM purchase_record pr
            JOIN upload_batch ub ON ub.id = pr.batch_id
            WHERE pr.stock_status = 'out_stock' AND ub.folder_stage = 'out_stock'
            ORDER BY updated_at ASC
            """
        ).fetchall()
    export_records_to_excel(IN_STOCK_MASTER_DIR / "库内总表.xlsx", in_stock_rows)
    export_records_to_excel(OUT_STOCK_MASTER_DIR / "出库历史总表.xlsx", out_stock_rows)


def export_records_to_excel(path: Path, rows: list[sqlite3.Row]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(
        [
            "批次编号",
            "表格行号",
            "采购日期",
            "物资名称",
            "规格型号",
            "单位",
            "数量",
            "单价",
            "金额",
            "供应商",
            "发票代码",
            "发票号码",
            "发票日期",
            "发票金额",
            "备注",
            "库存状态",
            "出库时间",
        ]
    )
    for row in rows:
        sheet.append(
            [
                row["batch_id"],
                row["row_no"],
                row["purchase_date"],
                row["item_name"],
                row["spec_model"],
                row["unit"],
                row["quantity"],
                row["unit_price"],
                row["amount"],
                row["supplier_name"],
                row["invoice_code"],
                row["invoice_number"],
                row["invoice_date"],
                row["invoice_amount"],
                row["remark"],
                row["stock_status"],
                row["reimbursed_at"],
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def build_template_workbook() -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "采购模板"
    sheet.append(list(HEADER_MAP.keys()))
    sheet.append(
        [
            "2026-06-18",
            "A4纸",
            "80g",
            "包",
            2,
            25,
            50,
            "文具店",
            "FP01",
            "INV-0001",
            "2026-06-18",
            50,
            "示例数据",
        ]
    )
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def export_reimbursement(record_ids: list[str], operator: str) -> tuple[bool, str]:
    if not record_ids:
        return False, "请选择需要出库的库内明细"
    with db_connection() as conn:
        placeholders = ",".join("?" for _ in record_ids)
        rows = conn.execute(
            f"""
            SELECT pr.* FROM purchase_record pr
            JOIN upload_batch ub ON ub.id = pr.batch_id
            WHERE pr.id IN ({placeholders}) AND pr.stock_status = 'in_stock' AND ub.folder_stage = 'in_stock'
            ORDER BY batch_id ASC, row_no ASC
            """,
            record_ids,
        ).fetchall()
        if not rows:
            return False, "没有可出库的库内明细"

        reimbursement_id = f"RB{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4]}"
        export_path = REIMBURSEMENT_EXPORT_DIR / f"{reimbursement_id}.xlsx"
        export_records_to_excel(export_path, rows)

        total_amount = sum(row["amount"] for row in rows)
        current_time = now_iso()
        conn.execute(
            """
            INSERT INTO reimbursement_batch (
                id, batch_name, extracted_by, extracted_at, record_count,
                total_amount, export_file_path, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                reimbursement_id,
                f"报销提取_{reimbursement_id}",
                operator,
                current_time,
                len(rows),
                total_amount,
                str(export_path),
                current_time,
            ),
        )

        affected_batches: set[str] = set()
        for row in rows:
            conn.execute(
                """
                INSERT INTO reimbursement_record (
                    id, reimbursement_batch_id, purchase_record_id,
                    invoice_number, reimbursed_at, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    uuid.uuid4().hex,
                    reimbursement_id,
                    row["id"],
                    row["invoice_number"],
                    current_time,
                    current_time,
                ),
            )
            conn.execute(
                """
                UPDATE purchase_record
                SET stock_status = 'out_stock', reimbursed_at = ?, reimbursement_batch_id = ?, updated_at = ?
                WHERE id = ?
                """,
                (current_time, reimbursement_id, current_time, row["id"]),
            )
            affected_batches.add(row["batch_id"])

        for batch_id in affected_batches:
            remaining = conn.execute(
                """
                SELECT COUNT(*) AS total FROM purchase_record
                WHERE batch_id = ? AND stock_status = 'in_stock'
                """,
                (batch_id,),
            ).fetchone()["total"]
            if remaining == 0:
                move_batch_folder(batch_id, "in_stock", "out_stock")
                conn.execute(
                    """
                    UPDATE upload_batch
                    SET status = 'out_stock', folder_stage = 'out_stock', form_file_path = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (str(batch_form_path(batch_id, "out_stock")), current_time, batch_id),
                )

    export_master_sheets()
    log_process("reimbursement", "info", f"生成出库批次 {reimbursement_id}")
    return True, reimbursement_id


def recent_logs(limit: int = 20) -> list[sqlite3.Row]:
    with db_connection() as conn:
        return conn.execute(
            """
            SELECT * FROM process_log
            ORDER BY created_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()


def dashboard_data() -> dict[str, Any]:
    with db_connection() as conn:
        counts = {
            "unregistered": conn.execute(
                "SELECT COUNT(*) AS total FROM upload_batch WHERE status = 'unregistered'"
            ).fetchone()["total"],
            "pending_review": conn.execute(
                "SELECT COUNT(*) AS total FROM upload_batch WHERE folder_stage = 'pending_review'"
            ).fetchone()["total"],
            "in_stock": conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM purchase_record pr
                JOIN upload_batch ub ON ub.id = pr.batch_id
                WHERE pr.stock_status = 'in_stock' AND ub.folder_stage = 'in_stock'
                """
            ).fetchone()["total"],
            "out_stock": conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM purchase_record pr
                JOIN upload_batch ub ON ub.id = pr.batch_id
                WHERE pr.stock_status = 'out_stock' AND ub.folder_stage = 'out_stock'
                """
            ).fetchone()["total"],
        }
        pending_batches = conn.execute(
            """
            SELECT
                ub.*,
                SUM(CASE WHEN spr.review_status = 'pending_review' THEN 1 ELSE 0 END) AS pending_rows,
                SUM(CASE WHEN spr.review_status = 'confirmed' THEN 1 ELSE 0 END) AS confirmed_rows,
                SUM(CASE WHEN spr.review_status = 'rejected' THEN 1 ELSE 0 END) AS rejected_rows
            FROM upload_batch ub
            LEFT JOIN staged_purchase_record spr ON spr.batch_id = ub.id
            WHERE ub.folder_stage = 'pending_review'
            GROUP BY ub.id
            ORDER BY ub.submitted_at ASC
            """
        ).fetchall()
        in_stock_rows = conn.execute(
            """
            SELECT pr.* FROM purchase_record pr
            JOIN upload_batch ub ON ub.id = pr.batch_id
            WHERE pr.stock_status = 'in_stock' AND ub.folder_stage = 'in_stock'
            ORDER BY pr.created_at ASC
            """
        ).fetchall()
        reimbursement_batches = conn.execute(
            """
            SELECT * FROM reimbursement_batch
            ORDER BY created_at DESC
            LIMIT 10
            """
        ).fetchall()
    return {
        "counts": counts,
        "pending_batches": pending_batches,
        "in_stock_rows": in_stock_rows,
        "reimbursement_batches": reimbursement_batches,
        "logs": recent_logs(),
    }


def is_admin_logged_in() -> bool:
    return bool(session.get("admin_authenticated"))


def require_admin():
    if is_admin_logged_in():
        return None
    flash("请输入管理员密码后再进入管理后台", "error")
    return redirect(url_for("admin_login"))


def get_batch(batch_id: str) -> sqlite3.Row | None:
    with db_connection() as conn:
        return conn.execute(
            "SELECT * FROM upload_batch WHERE id = ?",
            (batch_id,),
        ).fetchone()


def batch_rows(batch_id: str) -> list[sqlite3.Row]:
    with db_connection() as conn:
        return conn.execute(
            """
            SELECT * FROM staged_purchase_record
            WHERE batch_id = ?
            ORDER BY row_no ASC
            """,
            (batch_id,),
        ).fetchall()


class AgentThread:
    def __init__(self) -> None:
        self._stop_event = threading.Event()
        self._thread = threading.Thread(target=self._loop, daemon=True, name="material-agent")

    def start(self) -> None:
        if not self._thread.is_alive():
            self._thread.start()

    def stop(self) -> None:
        self._stop_event.set()

    def _loop(self) -> None:
        while not self._stop_event.is_set():
            try:
                run_agent_once()
            except Exception as exc:  # noqa: BLE001
                log_process("agent", "error", f"后台任务异常: {exc}")
            self._stop_event.wait(AGENT_INTERVAL_SECONDS)


app = Flask(__name__)
app.config["SECRET_KEY"] = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH

init_db()
agent_thread = AgentThread()
agent_thread.start()


@app.context_processor
def inject_helpers():
    return {
        "now_text": now_iso(),
        "admin_authenticated": is_admin_logged_in(),
    }


@app.route("/", methods=["GET", "POST"])
def upload_page():
    if request.method == "POST":
        team_name = normalize_text(request.form.get("team_name"))
        submitter_name = normalize_text(request.form.get("submitter_name"))
        submitter_id = normalize_text(request.form.get("submitter_id"))
        remark = normalize_text(request.form.get("remark"))
        form_file = request.files.get("form_file")

        if not team_name or not submitter_name or not submitter_id:
            flash("团队名称、提交人姓名、提交人编号不能为空", "error")
            return redirect(url_for("upload_page"))
        if form_file is None or not form_file.filename:
            flash("请上传采购表格", "error")
            return redirect(url_for("upload_page"))
        if Path(form_file.filename).suffix.lower() != ".xlsx":
            flash("只支持上传 xlsx 表格", "error")
            return redirect(url_for("upload_page"))

        batch_id = build_batch_id(submitter_id)
        batch_dir = UNREGISTERED_DIR / batch_id
        batch_dir.mkdir(parents=True, exist_ok=True)
        target_path = batch_dir / "form.xlsx"
        form_file.save(target_path)

        submitted_at = now_iso()
        meta = {
            "batch_id": batch_id,
            "team_name": team_name,
            "submitter_name": submitter_name,
            "submitter_id": submitter_id,
            "submitted_at": submitted_at,
            "form_file": "form.xlsx",
            "remark": remark,
            "folder_stage": "unregistered",
        }
        write_meta(batch_dir, meta)

        with db_connection() as conn:
            conn.execute(
                """
                INSERT INTO upload_batch (
                    id, team_name, submitter_name, submitter_id, submitted_at,
                    form_file_path, remark, status, folder_stage, review_note, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    team_name,
                    submitter_name,
                    submitter_id,
                    submitted_at,
                    str(target_path),
                    remark,
                    "unregistered",
                    "unregistered",
                    "",
                    submitted_at,
                    submitted_at,
                ),
            )
        log_process("upload", "info", "上传成功，等待 Agent 审核", batch_id)
        flash(f"上传成功，批次编号：{batch_id}", "success")
        return redirect(url_for("upload_page"))

    return render_template("upload.html")


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        password = normalize_text(request.form.get("password"))
        if password == ADMIN_PASSWORD:
            session["admin_authenticated"] = True
            flash("管理员登录成功", "success")
            return redirect(url_for("admin_dashboard"))
        flash("管理员密码错误", "error")
        return redirect(url_for("admin_login"))

    return render_template("admin_login.html")


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("admin_authenticated", None)
    flash("已退出管理后台", "success")
    return redirect(url_for("upload_page"))


@app.route("/admin")
def admin_dashboard():
    guard = require_admin()
    if guard is not None:
        return guard
    data = dashboard_data()
    return render_template("admin.html", **data)


@app.route("/admin/agent/run", methods=["POST"])
def admin_run_agent():
    guard = require_admin()
    if guard is not None:
        return guard
    processed = run_agent_once()
    flash(f"Agent 已执行，本次扫描批次数：{processed}", "success")
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/batches/<batch_id>")
def admin_batch_detail(batch_id: str):
    guard = require_admin()
    if guard is not None:
        return guard
    batch = get_batch(batch_id)
    if batch is None:
        flash("批次不存在", "error")
        return redirect(url_for("admin_dashboard"))
    rows = batch_rows(batch_id)
    return render_template("batch_detail.html", batch=batch, rows=rows)


@app.route("/admin/batches/<batch_id>/confirm-all", methods=["POST"])
def confirm_all(batch_id: str):
    guard = require_admin()
    if guard is not None:
        return guard
    rows = [row["id"] for row in batch_rows(batch_id) if row["review_status"] == "pending_review"]
    count = confirm_staged_rows(batch_id, rows, "管理员")
    flash(f"已确认 {count} 行入库", "success")
    return redirect(url_for("admin_batch_detail", batch_id=batch_id))


@app.route("/admin/batches/<batch_id>/confirm-selected", methods=["POST"])
def confirm_selected(batch_id: str):
    guard = require_admin()
    if guard is not None:
        return guard
    row_ids = request.form.getlist("row_ids")
    count = confirm_staged_rows(batch_id, row_ids, "管理员")
    flash(f"已确认 {count} 行入库", "success")
    return redirect(url_for("admin_batch_detail", batch_id=batch_id))


@app.route("/admin/batches/<batch_id>/reject-selected", methods=["POST"])
def reject_selected(batch_id: str):
    guard = require_admin()
    if guard is not None:
        return guard
    row_ids = request.form.getlist("row_ids")
    note = normalize_text(request.form.get("note")) or "人工复核未通过"
    count = reject_staged_rows(batch_id, row_ids, "管理员", note)
    flash(f"已打回 {count} 行", "success")
    return redirect(url_for("admin_batch_detail", batch_id=batch_id))


@app.route("/admin/batches/<batch_id>/complete", methods=["POST"])
def complete_batch_review(batch_id: str):
    guard = require_admin()
    if guard is not None:
        return guard
    success, message = finalize_batch_review(batch_id)
    flash(message, "success" if success else "error")
    return redirect(url_for("admin_batch_detail", batch_id=batch_id))


@app.route("/admin/reimburse", methods=["POST"])
def admin_reimburse():
    guard = require_admin()
    if guard is not None:
        return guard
    record_ids = request.form.getlist("record_ids")
    success, message = export_reimbursement(record_ids, "管理员")
    flash(f"出库完成，批次号：{message}" if success else message, "success" if success else "error")
    return redirect(url_for("admin_dashboard"))


@app.route("/downloads/form/<batch_id>")
def download_form(batch_id: str):
    guard = require_admin()
    if guard is not None:
        return guard
    batch = get_batch(batch_id)
    if batch is None:
        flash("批次不存在", "error")
        return redirect(url_for("admin_dashboard"))
    path = Path(batch["form_file_path"])
    if not path.exists():
        path = batch_form_path(batch_id, batch["folder_stage"])
    if not path.exists():
        flash("表格文件不存在", "error")
        return redirect(url_for("admin_dashboard"))
    return send_file(path, as_attachment=True, download_name=f"{batch_id}.xlsx")


@app.route("/downloads/template")
def download_template():
    stream = build_template_workbook()
    return send_file(
        stream,
        as_attachment=True,
        download_name="采购表格模板.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/downloads/reimbursement/<reimbursement_id>")
def download_reimbursement(reimbursement_id: str):
    guard = require_admin()
    if guard is not None:
        return guard
    with db_connection() as conn:
        batch = conn.execute(
            "SELECT * FROM reimbursement_batch WHERE id = ?",
            (reimbursement_id,),
        ).fetchone()
    if batch is None:
        flash("出库批次不存在", "error")
        return redirect(url_for("admin_dashboard"))
    path = Path(batch["export_file_path"])
    if not path.exists():
        flash("导出文件不存在", "error")
        return redirect(url_for("admin_dashboard"))
    return send_file(path, as_attachment=True, download_name=f"{reimbursement_id}.xlsx")


if __name__ == "__main__":
    app.run(host=APP_HOST, port=APP_PORT, debug=False)
