from __future__ import annotations

import base64
import csv
import hashlib
import hmac
import io
import json
import os
import shutil
import sqlite3
import threading
import time
import uuid
from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel


PROJECT_ROOT = Path(__file__).resolve().parents[2]
STORAGE_DIR = Path(os.environ.get("STORAGE_DIR", PROJECT_ROOT / "storage")).resolve()
DB_PATH = STORAGE_DIR / "system.db"

UNREGISTERED_DIR = STORAGE_DIR / "invoices" / "unregistered"
PENDING_REVIEW_DIR = STORAGE_DIR / "invoices" / "pending_review"
IN_STOCK_DIR = STORAGE_DIR / "invoices" / "in_stock"
OUT_STOCK_DIR = STORAGE_DIR / "invoices" / "out_stock"
PROCESSING_DIR = STORAGE_DIR / "invoices" / "processing"

DUPLICATE_ARCHIVE_DIR = STORAGE_DIR / "archive" / "duplicate_invoice"
PARSE_FAILED_DIR = STORAGE_DIR / "archive" / "parse_failed"
VALIDATION_FAILED_DIR = STORAGE_DIR / "archive" / "validation_failed"
REVIEW_REJECTED_DIR = STORAGE_DIR / "archive" / "review_rejected"

IN_STOCK_MASTER_DIR = STORAGE_DIR / "master" / "in_stock_master"
OUT_STOCK_MASTER_DIR = STORAGE_DIR / "master" / "out_stock_master"
REIMBURSEMENT_EXPORT_DIR = STORAGE_DIR / "master" / "reimbursement_export"
TEMP_DIR = STORAGE_DIR / "temp"
SITE_MEDIA_DIR = STORAGE_DIR / "site_media"

MAX_CONTENT_LENGTH = 60 * 1024 * 1024
MAX_VIDEO_CONTENT_LENGTH = 70 * 1024 * 1024
MAX_MEMBER_PHOTO_LENGTH = 8 * 1024 * 1024
AGENT_INTERVAL_SECONDS = int(os.getenv("AGENT_INTERVAL_SECONDS", "300"))
SECRET_KEY = os.getenv("SECRET_KEY", "material-agent-secret")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "wrprintk")
FRONTEND_ORIGIN = os.getenv("FRONTEND_ORIGIN", "http://127.0.0.1:3000,http://localhost:3000")
SITE_MODE = os.getenv("SITE_MODE", "full").strip().lower()
ENABLE_WRITE_API = os.getenv("ENABLE_WRITE_API", "true").strip().lower() in {"1", "true", "yes", "on"}
AGENT_LOCK = threading.Lock()
FORUM_PUBLIC_STATUS = "approved"
FORUM_CONTENT_STATUSES = {"pending", "approved", "rejected", "hidden"}
SSL_APPLICATION_STATUSES = {"pending", "approved", "rejected"}
SSL_INTERVIEW_DIRECTIONS = {"机械", "电控", "硬件", "算法", "视觉", "运营"}
MARKET_ITEM_STATUSES = {"available", "sold", "delisted"}
HOME_ASSET_KINDS = {"video", "image"}
HOME_VIDEO_MIME_TYPES = {"video/mp4", "video/webm", "video/quicktime"}
HOME_IMAGE_MIME_TYPES = {"image/jpeg", "image/png", "image/webp", "image/gif"}
MEMBER_PHOTO_MIME_TYPES = {"image/jpeg", "image/png", "image/webp"}
DANMAKU_TRACKS = 7
DANMAKU_COLORS = ["#ffffff"]

HEADER_MAP = {
    "采购日期": "purchase_date",
    "物资名称": "item_name",
    "规格型号": "spec_model",
    "单位": "unit",
    "数量": "quantity",
    "单价": "unit_price",
    "金额": "amount",
    "供应商": "supplier_name",
    "发票代码": "invoice_code",
    "发票号码": "invoice_number",
    "发票日期": "invoice_date",
    "发票金额": "invoice_amount",
    "备注": "remark",
}
REQUIRED_HEADERS = ["采购日期", "物资名称", "单位", "数量", "单价", "金额", "发票号码"]


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value: Any) -> str:
    return "".join(normalize_text(value).lstrip("\ufeff").split())


def build_batch_id(submitter_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    safe_submitter = "".join(ch for ch in submitter_id if ch.isalnum() or ch in ("-", "_")) or "member"
    return f"{timestamp}_{safe_submitter}_{uuid.uuid4().hex[:4]}"


def normalize_number(value: Any, field_name: str, required: bool) -> float | None:
    text = normalize_text(value)
    if not text:
        if required:
            raise ValueError(f"{field_name}不能为空")
        return None
    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field_name}格式错误") from exc


def normalize_date(value: Any, field_name: str, required: bool) -> str:
    if value in (None, ""):
        if required:
            raise ValueError(f"{field_name}不能为空")
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_text(value).replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"{field_name}格式错误")


def ensure_directories() -> None:
    for path in [
        UNREGISTERED_DIR,
        PENDING_REVIEW_DIR,
        IN_STOCK_DIR,
        OUT_STOCK_DIR,
        PROCESSING_DIR,
        DUPLICATE_ARCHIVE_DIR,
        PARSE_FAILED_DIR,
        VALIDATION_FAILED_DIR,
        REVIEW_REJECTED_DIR,
        IN_STOCK_MASTER_DIR,
        OUT_STOCK_MASTER_DIR,
        REIMBURSEMENT_EXPORT_DIR,
        TEMP_DIR,
        SITE_MEDIA_DIR,
    ]:
        path.mkdir(parents=True, exist_ok=True)


@contextmanager
def db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db() -> None:
    ensure_directories()
    with db_connection() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS upload_batch (
                id TEXT PRIMARY KEY,
                team_name TEXT NOT NULL,
                submitter_name TEXT NOT NULL,
                submitter_id TEXT NOT NULL,
                submitted_at TEXT NOT NULL,
                form_file_path TEXT NOT NULL,
                remark TEXT DEFAULT '',
                status TEXT NOT NULL,
                folder_stage TEXT NOT NULL,
                review_note TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS staged_purchase_record (
                id TEXT PRIMARY KEY,
                batch_id TEXT NOT NULL,
                row_no INTEGER NOT NULL,
                purchase_date TEXT NOT NULL,
                item_name TEXT NOT NULL,
                spec_model TEXT DEFAULT '',
                unit TEXT NOT NULL,
                quantity REAL NOT NULL,
                unit_price REAL NOT NULL,
                amount REAL NOT NULL,
                supplier_name TEXT DEFAULT '',
                invoice_code TEXT DEFAULT '',
                invoice_number TEXT NOT NULL,
                invoice_date TEXT DEFAULT '',
                invoice_amount REAL,
                remark TEXT DEFAULT '',
                review_status TEXT NOT NULL,
                review_note TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS purchase_record (
                id TEXT PRIMARY KEY,
                batch_id TEXT NOT NULL,
                staged_record_id TEXT NOT NULL UNIQUE,
                row_no INTEGER NOT NULL,
                purchase_date TEXT NOT NULL,
                item_name TEXT NOT NULL,
                spec_model TEXT DEFAULT '',
                unit TEXT NOT NULL,
                quantity REAL NOT NULL,
                unit_price REAL NOT NULL,
                amount REAL NOT NULL,
                supplier_name TEXT DEFAULT '',
                invoice_code TEXT DEFAULT '',
                invoice_number TEXT NOT NULL,
                invoice_date TEXT DEFAULT '',
                invoice_amount REAL,
                remark TEXT DEFAULT '',
                stock_status TEXT NOT NULL,
                reimbursed_at TEXT DEFAULT '',
                reimbursement_batch_id TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS invoice_registry (
                id TEXT PRIMARY KEY,
                invoice_number TEXT NOT NULL UNIQUE,
                batch_id TEXT NOT NULL,
                purchase_record_id TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS reimbursement_batch (
                id TEXT PRIMARY KEY,
                batch_name TEXT NOT NULL,
                extracted_by TEXT NOT NULL,
                extracted_at TEXT NOT NULL,
                record_count INTEGER NOT NULL,
                total_amount REAL NOT NULL,
                export_file_path TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS reimbursement_record (
                id TEXT PRIMARY KEY,
                reimbursement_batch_id TEXT NOT NULL,
                purchase_record_id TEXT NOT NULL,
                invoice_number TEXT NOT NULL,
                reimbursed_at TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS process_log (
                id TEXT PRIMARY KEY,
                batch_id TEXT,
                stage TEXT NOT NULL,
                level TEXT NOT NULL,
                message TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS season_plan (
                id TEXT PRIMARY KEY,
                season_year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                group_name TEXT DEFAULT '',
                robot_type TEXT NOT NULL,
                task_title TEXT NOT NULL,
                status TEXT NOT NULL,
                target TEXT NOT NULL,
                assignee_account TEXT DEFAULT '',
                is_completed INTEGER NOT NULL DEFAULT 0,
                display_order INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS site_account (
                account TEXT PRIMARY KEY,
                password_hash TEXT NOT NULL,
                full_name TEXT DEFAULT '',
                gender TEXT DEFAULT '',
                grade TEXT DEFAULT '',
                member_status TEXT DEFAULT '',
                permission_level TEXT DEFAULT '',
                department TEXT DEFAULT '',
                cohort TEXT DEFAULT '',
                role TEXT DEFAULT '',
                phone TEXT DEFAULT '',
                email TEXT DEFAULT '',
                bio TEXT DEFAULT '',
                photo_url TEXT DEFAULT '',
                reward_score INTEGER NOT NULL DEFAULT 0,
                image2_allowed INTEGER NOT NULL DEFAULT 0,
                is_disabled INTEGER NOT NULL DEFAULT 0,
                admin_note TEXT DEFAULT '',
                last_login_at TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS site_account_admin_log (
                id TEXT PRIMARY KEY,
                account TEXT NOT NULL,
                action TEXT NOT NULL,
                detail TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY (account) REFERENCES site_account(account)
            );

            CREATE TABLE IF NOT EXISTS ssl_interview_application (
                id TEXT PRIMARY KEY,
                applicant_account TEXT NOT NULL UNIQUE,
                self_intro TEXT NOT NULL,
                interview_direction TEXT NOT NULL,
                interview_time TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                interview_location TEXT DEFAULT '',
                rejection_reason TEXT DEFAULT '',
                reviewed_by TEXT DEFAULT '',
                reviewed_at TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (applicant_account) REFERENCES site_account(account)
            );

            CREATE TABLE IF NOT EXISTS site_message (
                id TEXT PRIMARY KEY,
                recipient_account TEXT NOT NULL,
                category TEXT NOT NULL,
                title TEXT NOT NULL,
                content TEXT NOT NULL,
                related_id TEXT DEFAULT '',
                read_at TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY (recipient_account) REFERENCES site_account(account)
            );

            CREATE TABLE IF NOT EXISTS forum_post (
                id TEXT PRIMARY KEY,
                title TEXT NOT NULL,
                content TEXT NOT NULL,
                author_account TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'approved',
                reject_reason TEXT DEFAULT '',
                reviewed_by TEXT DEFAULT '',
                reviewed_at TEXT DEFAULT '',
                deleted_at TEXT DEFAULT '',
                is_pinned INTEGER NOT NULL DEFAULT 0,
                is_locked INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (author_account) REFERENCES site_account(account)
            );

            CREATE TABLE IF NOT EXISTS forum_reply (
                id TEXT PRIMARY KEY,
                post_id TEXT NOT NULL,
                content TEXT NOT NULL,
                author_account TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'approved',
                reject_reason TEXT DEFAULT '',
                reviewed_by TEXT DEFAULT '',
                reviewed_at TEXT DEFAULT '',
                deleted_at TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (post_id) REFERENCES forum_post(id),
                FOREIGN KEY (author_account) REFERENCES site_account(account)
            );

            CREATE TABLE IF NOT EXISTS flea_market_item (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                image_src TEXT DEFAULT '',
                image_alt TEXT DEFAULT '',
                author_account TEXT NOT NULL,
                team TEXT DEFAULT '',
                location TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'available',
                summary TEXT NOT NULL,
                detail TEXT NOT NULL,
                contact TEXT NOT NULL,
                tags TEXT DEFAULT '[]',
                delisted_at TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (author_account) REFERENCES site_account(account)
            );

            CREATE TABLE IF NOT EXISTS homepage_asset (
                id TEXT PRIMARY KEY,
                kind TEXT NOT NULL,
                url TEXT NOT NULL,
                storage_path TEXT DEFAULT '',
                original_filename TEXT DEFAULT '',
                mime_type TEXT DEFAULT '',
                size_bytes INTEGER NOT NULL DEFAULT 0,
                alt TEXT DEFAULT '',
                display_order INTEGER NOT NULL DEFAULT 0,
                is_enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS homepage_quote (
                id TEXT PRIMARY KEY,
                text TEXT NOT NULL,
                source TEXT DEFAULT '',
                display_order INTEGER NOT NULL DEFAULT 0,
                is_enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS homepage_recruitment_banner (
                id TEXT PRIMARY KEY,
                text TEXT NOT NULL,
                action_text TEXT NOT NULL,
                is_enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS homepage_profile (
                id TEXT PRIMARY KEY,
                team_name TEXT NOT NULL,
                team_intro TEXT NOT NULL,
                stats_json TEXT NOT NULL DEFAULT '[]',
                awards_json TEXT NOT NULL DEFAULT '[]',
                recruitment_json TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS homepage_danmaku (
                id TEXT PRIMARY KEY,
                image_key TEXT DEFAULT '',
                image_src TEXT NOT NULL,
                author_account TEXT DEFAULT '',
                author_name TEXT DEFAULT '',
                text TEXT NOT NULL,
                track INTEGER NOT NULL,
                color TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                reviewed_by TEXT DEFAULT '',
                reviewed_at TEXT DEFAULT '',
                created_at_ms INTEGER NOT NULL,
                duration REAL NOT NULL,
                delay REAL NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS recruitment_question (
                id TEXT PRIMARY KEY,
                author_account TEXT NOT NULL,
                content TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (author_account) REFERENCES site_account(account)
            );

            CREATE TABLE IF NOT EXISTS recruitment_faq (
                id TEXT PRIMARY KEY,
                question TEXT NOT NULL,
                answer TEXT NOT NULL,
                display_order INTEGER NOT NULL DEFAULT 0,
                is_enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_season_plan_period
            ON season_plan (season_year, month, display_order);

            CREATE INDEX IF NOT EXISTS idx_forum_post_created_at
            ON forum_post (created_at DESC);

            CREATE INDEX IF NOT EXISTS idx_forum_reply_post_created_at
            ON forum_reply (post_id, created_at);

            """
        )
        ensure_season_plan_schema(conn)
        ensure_site_account_profile_columns(conn)
        ensure_site_message_read_column(conn)
        ensure_homepage_profile_columns(conn)
        ensure_homepage_danmaku_columns(conn)
        ensure_forum_moderation_columns(conn)
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_forum_post_status_created_at
            ON forum_post (status, deleted_at, is_pinned, created_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_forum_reply_status_created_at
            ON forum_reply (status, deleted_at, created_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_forum_post_author_status_updated
            ON forum_post (author_account, status, deleted_at, updated_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_forum_reply_author_status_updated
            ON forum_reply (author_account, status, deleted_at, updated_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_site_account_member_permission
            ON site_account (member_status, permission_level, image2_allowed, is_disabled)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_site_account_reward_ranking
            ON site_account (member_status, is_disabled, reward_score DESC, updated_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_site_account_public_members
            ON site_account (is_disabled, member_status, cohort, department, updated_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_site_account_admin_log_account
            ON site_account_admin_log (account, created_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_ssl_interview_status_time
            ON ssl_interview_application (status, interview_time, created_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_site_message_recipient_created
            ON site_message (recipient_account, created_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_site_message_unread
            ON site_message (recipient_account, read_at, created_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_flea_market_public
            ON flea_market_item (status, delisted_at, created_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_flea_market_author
            ON flea_market_item (author_account, updated_at DESC)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_homepage_asset_kind_order
            ON homepage_asset (kind, is_enabled, display_order, created_at)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_homepage_quote_order
            ON homepage_quote (is_enabled, display_order, created_at)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_homepage_danmaku_image_created
            ON homepage_danmaku (image_key, created_at_ms)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_homepage_danmaku_status_created
            ON homepage_danmaku (status, created_at_ms DESC)
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_recruitment_question_created ON recruitment_question (created_at DESC)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_recruitment_faq_order ON recruitment_faq (is_enabled, display_order, created_at)")
        seed_season_plan(conn)
        seed_homepage_content(conn)


def row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


def seed_season_plan(conn: sqlite3.Connection) -> None:
    existing = conn.execute(
        "SELECT COUNT(*) AS total FROM season_plan WHERE season_year = ? AND month = ?",
        (2026, 6),
    ).fetchone()
    if existing and existing["total"] > 0:
        return
    plans = [
        ("英雄兵种", "云台与发射联调", "准备中", "完成云台通信、发射链路和赛前检查清单。"),
        ("步兵兵种", "底盘功率控制", "准备中", "完成底盘通信、功率管理和基础控制联调。"),
        ("工程兵种", "机构方案验证", "准备中", "完成关键机构方案评审和第一轮装配验证。"),
        ("哨兵兵种", "自动导航测试", "准备中", "完成自动导航、策略接口和仿真数据整理。"),
    ]
    timestamp = now_iso()
    conn.executemany(
        """
        INSERT INTO season_plan (
            id, season_year, month, group_name, robot_type, task_title, status, target,
            assignee_account, is_completed, display_order, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, '', 0, ?, ?, ?)
        """,
        [
            (uuid.uuid4().hex, 2026, 6, f"{robot_type}：{task_title}", robot_type, task_title, status, target, index, timestamp, timestamp)
            for index, (robot_type, task_title, status, target) in enumerate(plans, start=1)
        ],
    )


HOME_IMAGE_SEEDS = [
    ("/home-carousel/team-01.jpeg", "PRINTK 成员在实验室演示康复机器人设备"),
    ("/home-carousel/team-02.jpeg", "PRINTK 队员在赛事现场交流"),
    ("/home-carousel/team-03.jpeg", "PRINTK 队员在场馆内讨论比赛细节"),
    ("/home-carousel/team-04.jpeg", "PRINTK 成员在赛场调试机器人"),
    ("/home-carousel/team-05.jpeg", "PRINTK 队员在比赛现场观察机器人状态"),
    ("/home-carousel/team-06.jpeg", "PRINTK 队员在场边关注比赛进程"),
    ("/home-carousel/team-07.jpeg", "PRINTK 成员在场馆通道集合"),
    ("/home-carousel/team-08.png", "PRINTK 战队赛季全员合影"),
    ("/home-carousel/team-09.jpg", "PRINTK 战队在 RoboMaster 现场合影"),
    ("/home-carousel/team-10.jpg", "PRINTK 成员围绕机器人开展线下交流"),
    ("/home-carousel/team-11.jpg", "PRINTK 队员围绕电脑集中讨论调试方案"),
    ("/home-carousel/team-12.jpeg", "PRINTK 战队与机器人在赛场内合影留念"),
    ("/home-carousel/team-13.jpeg", "PRINTK 队员在比赛现场近距离调试机器人"),
]

HOME_QUOTE_SEEDS = [
    ("道阻且长，行则将至", "PRINTK 赛季口号"),
    ("为青春赋予荣耀，让思考拥有力量", "RoboMaster 赛事理念"),
    ("服务全球青年工程师成为追求极致、有实干精神的梦想家", "RoboMaster 高校系列赛"),
    ("崇尚科学与创新，擅于反思，勇于实践，热爱分享", "RoboMaster 赛事理念"),
    ("初心高于胜负，每一份努力都值得被肯定", "RoboMaster 组织奖文化"),
    ("以学术价值为根基，培养具有工程思维、拥有实干精神的综合素质人才", "RoboMaster 赛事愿景"),
    ("勇于创新、追求极致、崇尚实干、具备视野和远见", "RoboMaster 专属招聘通道"),
]

HOME_PROFILE_TEAM_NAME = "PRINTK 机甲大师战队"
HOME_PROFILE_TEAM_INTRO = (
    "PRINTK 机甲大师战队成立于 2024 年秋季，基地位于贵州大学明正楼科技园 1 楼报告厅，"
    "现有正式队员 30 余人。战队曾获 2025 赛季高校联盟赛广西站步兵对抗赛季军，并在 "
    "2026 赛季高校联盟赛重庆站首次完整出征步兵对抗赛、工程挑战赛与 3v3 对抗赛三个赛项。"
)
HOME_PROFILE_STATS = [
    {"value": "4", "label": "核心组别"},
    {"value": "7", "label": "兵种方向"},
    {"value": "2026", "label": "赛季规划"},
]
HOME_PROFILE_AWARDS = [
    {"title": "RoboMaster 赛事奖项", "meta": "奖状图片占位", "image_url": "", "image_alt": "", "display_order": 1},
    {"title": "赛季工程成果", "meta": "奖杯图片占位", "image_url": "", "image_alt": "", "display_order": 2},
    {"title": "校级竞赛荣誉", "meta": "证书图片占位", "image_url": "", "image_alt": "", "display_order": 3},
    {"title": "技术创新成果", "meta": "奖项图片占位", "image_url": "", "image_alt": "", "display_order": 4},
    {"title": "团队建设荣誉", "meta": "合影图片占位", "image_url": "", "image_alt": "", "display_order": 5},
    {"title": "年度贡献奖项", "meta": "荣誉图片占位", "image_url": "", "image_alt": "", "display_order": 6},
]
HOME_RECRUITMENT_CONTENT = {
    "season_label": "2028 赛季招新",
    "title": "加入 PRINTK，\n把热爱做成能上场的机器人",
    "intro": "从赛事认知到分组实践，找到适合自己的方向，和队友一起把想法做成真正能上场的机器人。",
    "event_kicker": "01 / ABOUT THE EVENT",
    "event_title": "RoboMaster 机甲大师赛事介绍",
    "event_description": "RoboMaster 机甲大师赛是国内顶尖大学生工科竞技赛事，被誉为青年工程师的培养摇篮。赛事主要分为机甲对抗赛、人工智能挑战赛、单项技能赛等多个赛道，综合考验机械结构设计、电控编程、机器视觉算法与团队运营能力。",
    "events": [
        {
            "id": "robomaster",
            "name": "RoboMaster",
            "kicker": "01 / 机甲大师",
            "title": "RoboMaster 赛事介绍",
            "description": "RoboMaster 机甲大师赛聚焦大学生机器人对抗，综合考验机械结构、电控编程、机器视觉与团队协作能力。",
        },
        {
            "id": "robocon",
            "name": "Robocon",
            "kicker": "02 / 亚太大学生机器人大赛",
            "title": "Robocon 赛事介绍",
            "description": "Robocon 鼓励大学生围绕年度主题自主设计和制作机器人，在限定场地完成任务挑战，强调创意、工程实现与现场协作。",
        },
        {
            "id": "robocup",
            "name": "RoboCup",
            "kicker": "03 / 世界机器人足球赛",
            "title": "RoboCup 赛事介绍",
            "description": "RoboCup 以全自主机器人足球为核心，训练感知、定位、路径规划和多机协同，让算法决策在真实赛场中持续进化。",
        },
    ],
    "groups_kicker": "02 / JOIN PRINTK",
    "groups_title": "PRINTK 五大组别",
    "groups": [
        {"name": "机械组", "summary": "负责结构设计、加工装配与整机维护。"},
        {"name": "电控组", "summary": "负责电气系统、嵌入式控制与整车联调。"},
        {"name": "硬件组", "summary": "负责电路板、传感器和硬件链路验证。"},
        {"name": "算法组", "summary": "负责视觉识别、运动控制与数据复盘。"},
        {"name": "运营组", "summary": "负责赛事运营、宣传内容与团队协作。"},
    ],
    "qr_text": "扫码进群即可报名咨询",
}


def seed_homepage_content(conn: sqlite3.Connection) -> None:
    timestamp = now_iso()
    profile_count = conn.execute("SELECT COUNT(*) AS total FROM homepage_profile").fetchone()["total"]
    if profile_count == 0:
        conn.execute(
            """
            INSERT INTO homepage_profile (
                id, team_name, team_intro, stats_json, awards_json, recruitment_json, created_at, updated_at
            )
            VALUES ('profile', ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                HOME_PROFILE_TEAM_NAME,
                HOME_PROFILE_TEAM_INTRO,
                json.dumps(HOME_PROFILE_STATS, ensure_ascii=False),
                json.dumps(HOME_PROFILE_AWARDS, ensure_ascii=False),
                json.dumps(HOME_RECRUITMENT_CONTENT, ensure_ascii=False),
                timestamp,
                timestamp,
            ),
        )
    conn.executemany(
        """
        INSERT OR IGNORE INTO homepage_recruitment_banner (
            id, text, action_text, is_enabled, created_at, updated_at
        )
        VALUES (?, ?, ?, 1, ?, ?)
        """,
        (
            ("recruitment", "2027赛季招新中", "点击跳转", timestamp, timestamp),
            ("campus-competition", "2027贵州大学机甲大师校内赛", "由此报名", timestamp, timestamp),
        ),
    )
    video_count = conn.execute("SELECT COUNT(*) AS total FROM homepage_asset WHERE kind = 'video'").fetchone()["total"]
    if video_count == 0:
        conn.execute(
            """
            INSERT INTO homepage_asset (
                id, kind, url, original_filename, mime_type, size_bytes, alt,
                display_order, is_enabled, created_at, updated_at
            )
            VALUES (?, 'video', ?, ?, 'video/mp4', ?, ?, 1, 1, ?, ?)
            """,
            (uuid.uuid4().hex, "/season-promo.mp4", "欢送老登之夜.mp4", 6233758, "赛季宣传视频", timestamp, timestamp),
        )

    image_count = conn.execute("SELECT COUNT(*) AS total FROM homepage_asset WHERE kind = 'image'").fetchone()["total"]
    if image_count == 0:
        conn.executemany(
            """
            INSERT INTO homepage_asset (
                id, kind, url, original_filename, mime_type, size_bytes, alt,
                display_order, is_enabled, created_at, updated_at
            )
            VALUES (?, 'image', ?, ?, '', 0, ?, ?, 1, ?, ?)
            """,
            [
                (uuid.uuid4().hex, url, Path(url).name, alt, index, timestamp, timestamp)
                for index, (url, alt) in enumerate(HOME_IMAGE_SEEDS, start=1)
            ],
        )

    quote_count = conn.execute("SELECT COUNT(*) AS total FROM homepage_quote").fetchone()["total"]
    if quote_count == 0:
        conn.executemany(
            """
            INSERT INTO homepage_quote (
                id, text, source, display_order, is_enabled, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, 1, ?, ?)
            """,
            [
                (uuid.uuid4().hex, text, source, index, timestamp, timestamp)
                for index, (text, source) in enumerate(HOME_QUOTE_SEEDS, start=1)
            ],
        )


def make_token(role: str) -> str:
    payload = f"{role}:{int(time.time())}"
    signature = hmac.new(SECRET_KEY.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    raw = f"{payload}:{signature}".encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("utf-8")


def normalize_account(value: str) -> str:
    account = value.strip()
    for _ in range(3):
        decoded = unquote(account)
        if decoded == account:
            break
        account = decoded
    if not account or len(account) > 32:
        raise HTTPException(status_code=400, detail="账号长度需为 1 到 32 个字符")
    return account


def ensure_season_plan_schema(conn: sqlite3.Connection) -> None:
    existing_columns = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(season_plan)").fetchall()
    }
    column_definitions = {
        "group_name": "TEXT DEFAULT ''",
        "robot_type": "TEXT NOT NULL DEFAULT ''",
        "task_title": "TEXT NOT NULL DEFAULT ''",
        "assignee_account": "TEXT DEFAULT ''",
        "is_completed": "INTEGER NOT NULL DEFAULT 0",
    }
    for column, definition in column_definitions.items():
        if column not in existing_columns:
            conn.execute(f"ALTER TABLE season_plan ADD COLUMN {column} {definition}")

    if "group_name" in existing_columns:
        conn.execute(
            """
            UPDATE season_plan
            SET robot_type = COALESCE(NULLIF(robot_type, ''), group_name),
                task_title = COALESCE(NULLIF(task_title, ''), target)
            WHERE COALESCE(robot_type, '') = '' OR COALESCE(task_title, '') = ''
            """
        )


def ensure_site_account_profile_columns(conn: sqlite3.Connection) -> None:
    existing_columns = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(site_account)").fetchall()
    }
    profile_columns = {
        "full_name": "TEXT DEFAULT ''",
        "gender": "TEXT DEFAULT ''",
        "grade": "TEXT DEFAULT ''",
        "member_status": "TEXT DEFAULT ''",
        "permission_level": "TEXT DEFAULT ''",
        "department": "TEXT DEFAULT ''",
        "cohort": "TEXT DEFAULT ''",
        "role": "TEXT DEFAULT ''",
        "phone": "TEXT DEFAULT ''",
        "email": "TEXT DEFAULT ''",
        "bio": "TEXT DEFAULT ''",
        "photo_url": "TEXT DEFAULT ''",
        "reward_score": "INTEGER NOT NULL DEFAULT 0",
        "image2_allowed": "INTEGER NOT NULL DEFAULT 0",
        "is_disabled": "INTEGER NOT NULL DEFAULT 0",
        "admin_note": "TEXT DEFAULT ''",
        "last_login_at": "TEXT DEFAULT ''",
    }
    for column, definition in profile_columns.items():
        if column not in existing_columns:
            conn.execute(f"ALTER TABLE site_account ADD COLUMN {column} {definition}")
    conn.execute(
        """
        UPDATE site_account
        SET permission_level = CASE member_status
            WHEN '管理员' THEN '管理'
            WHEN '队长' THEN '队长'
            WHEN '兵种组长' THEN '兵种组长'
            ELSE COALESCE(NULLIF(permission_level, ''), '普通队员')
        END
        WHERE COALESCE(permission_level, '') = ''
            OR member_status IN ('管理员', '队长', '兵种组长')
        """
    )
    conn.execute(
        """
        UPDATE site_account
        SET member_status = '正式队员'
        WHERE member_status IN ('管理员', '队长', '兵种组长')
        """
    )
    conn.execute(
        """
        UPDATE site_account
        SET reward_score = 0
        WHERE member_status NOT IN ('正式队员', '老队员')
            AND COALESCE(reward_score, 0) <> 0
        """
    )


def ensure_site_message_read_column(conn: sqlite3.Connection) -> None:
    existing_columns = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(site_message)").fetchall()
    }
    if "read_at" not in existing_columns:
        conn.execute("ALTER TABLE site_message ADD COLUMN read_at TEXT DEFAULT ''")


def ensure_forum_moderation_columns(conn: sqlite3.Connection) -> None:
    post_columns = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(forum_post)").fetchall()
    }
    post_definitions = {
        "status": "TEXT NOT NULL DEFAULT 'approved'",
        "reject_reason": "TEXT DEFAULT ''",
        "reviewed_by": "TEXT DEFAULT ''",
        "reviewed_at": "TEXT DEFAULT ''",
        "deleted_at": "TEXT DEFAULT ''",
        "is_pinned": "INTEGER NOT NULL DEFAULT 0",
        "is_locked": "INTEGER NOT NULL DEFAULT 0",
    }
    for column, definition in post_definitions.items():
        if column not in post_columns:
            conn.execute(f"ALTER TABLE forum_post ADD COLUMN {column} {definition}")

    reply_columns = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(forum_reply)").fetchall()
    }
    reply_definitions = {
        "status": "TEXT NOT NULL DEFAULT 'approved'",
        "reject_reason": "TEXT DEFAULT ''",
        "reviewed_by": "TEXT DEFAULT ''",
        "reviewed_at": "TEXT DEFAULT ''",
        "deleted_at": "TEXT DEFAULT ''",
        "updated_at": "TEXT NOT NULL DEFAULT ''",
    }
    for column, definition in reply_definitions.items():
        if column not in reply_columns:
            conn.execute(f"ALTER TABLE forum_reply ADD COLUMN {column} {definition}")
    conn.execute(
        """
        UPDATE forum_reply
        SET updated_at = created_at
        WHERE COALESCE(updated_at, '') = ''
        """
    )


def ensure_homepage_profile_columns(conn: sqlite3.Connection) -> None:
    existing_columns = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(homepage_profile)").fetchall()
    }
    if "recruitment_json" not in existing_columns:
        conn.execute("ALTER TABLE homepage_profile ADD COLUMN recruitment_json TEXT NOT NULL DEFAULT '{}'")


def ensure_homepage_danmaku_columns(conn: sqlite3.Connection) -> None:
    existing_columns = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(homepage_danmaku)").fetchall()
    }
    column_definitions = {
        "image_key": "TEXT DEFAULT ''",
        "author_account": "TEXT DEFAULT ''",
        "author_name": "TEXT DEFAULT ''",
        "status": "TEXT NOT NULL DEFAULT 'approved'",
        "reviewed_by": "TEXT DEFAULT ''",
        "reviewed_at": "TEXT DEFAULT ''",
    }
    for column, definition in column_definitions.items():
        if column not in existing_columns:
            conn.execute(f"ALTER TABLE homepage_danmaku ADD COLUMN {column} {definition}")

    conn.execute(
        """
        UPDATE homepage_danmaku
        SET image_key = image_src
        WHERE COALESCE(image_key, '') = ''
        """
    )

    conn.execute(
        """
        UPDATE homepage_danmaku
        SET author_name = COALESCE(
            NULLIF(author_name, ''),
            NULLIF((SELECT full_name FROM site_account WHERE site_account.account = homepage_danmaku.author_account), ''),
            author_account,
            ''
        )
        WHERE COALESCE(author_name, '') = ''
        """
    )


GENDER_OPTIONS = {"", "男", "女", "其他"}
GRADE_OPTIONS = {"", "大一", "大二", "大三", "大四", "研究生"}
PLAN_EDITOR_PERMISSION_OPTIONS = {"兵种组长", "部门组长", "队长", "管理"}
MEMBER_STATUS_OPTIONS = {"", "非战队队员", "梯队队员", "正式队员", "老队员", "退役队员", "老师"}
PERMISSION_LEVEL_OPTIONS = {"", "普通队员", "兵种组长", "部门组长", "队长", "管理"}
REWARD_STATUS_OPTIONS = {"正式队员", "老队员"}
DEPARTMENT_OPTIONS = {"", "队长", "项管", "机械组", "电控组", "硬件组", "算法组", "运营组", "电控", "机械", "算法", "运营"}
SEASON_PLAN_ROBOT_TYPES = {"英雄兵种", "步兵兵种", "工程兵种", "哨兵兵种"}


def normalize_limited_text(value: str, field_name: str, max_length: int = 80) -> str:
    text = value.strip()
    if len(text) > max_length:
        raise HTTPException(status_code=400, detail=f"{field_name}不能超过 {max_length} 个字符")
    return text


def normalize_choice(value: str, field_name: str, allowed_values: set[str]) -> str:
    text = value.strip()
    if text not in allowed_values:
        raise HTTPException(status_code=400, detail=f"{field_name}格式错误")
    return text


def normalize_photo_url(value: str) -> str:
    text = normalize_limited_text(value, "成员照片", 300)
    if text and not (text.startswith("/") or text.startswith("http://") or text.startswith("https://")):
        raise HTTPException(status_code=400, detail="成员照片必须是 / 开头路径或 http(s) 链接")
    return text


def normalize_site_profile(profile: "SiteAccountProfile") -> dict[str, str]:
    return {
        "full_name": normalize_limited_text(profile.full_name, "姓名", 32),
        "gender": normalize_choice(profile.gender, "性别", GENDER_OPTIONS),
        "grade": normalize_choice(profile.grade, "年级", GRADE_OPTIONS),
        "member_status": normalize_choice(profile.member_status, "身份信息", MEMBER_STATUS_OPTIONS),
        "permission_level": normalize_choice(profile.permission_level, "权限", PERMISSION_LEVEL_OPTIONS),
        "department": normalize_choice(profile.department, "部门信息", DEPARTMENT_OPTIONS),
        "cohort": normalize_limited_text(profile.cohort, "届别", 32),
        "role": normalize_limited_text(profile.role, "职责", 80),
        "phone": normalize_limited_text(profile.phone, "联系电话", 32),
        "email": normalize_limited_text(profile.email, "邮箱", 80),
        "bio": normalize_limited_text(profile.bio, "个人说明", 200),
        "photo_url": normalize_photo_url(profile.photo_url),
    }


def reward_eligible(member_status: str) -> bool:
    return member_status in REWARD_STATUS_OPTIONS


def normalize_reward_score(value: int) -> int:
    if value < 0:
        raise HTTPException(status_code=400, detail="奖励分不能小于 0")
    if value > 999999:
        raise HTTPException(status_code=400, detail="奖励分不能超过 999999")
    return value


def site_account_response(row: sqlite3.Row, include_admin: bool = False) -> dict[str, Any]:
    eligible = reward_eligible(row["member_status"] or "")
    data = {
        "account": row["account"],
        "full_name": row["full_name"] or "",
        "gender": row["gender"] or "",
        "grade": row["grade"] or "",
        "member_status": row["member_status"] or "",
        "permission_level": row["permission_level"] or "",
        "department": row["department"] or "",
        "cohort": row["cohort"] or "",
        "role": row["role"] or "",
        "phone": row["phone"] or "",
        "email": row["email"] or "",
        "bio": row["bio"] or "",
        "photo_url": row["photo_url"] or "",
        "reward_score": row["reward_score"] if eligible else 0,
        "reward_eligible": eligible,
        "image2_allowed": bool(row["image2_allowed"]),
        "is_disabled": bool(row["is_disabled"]),
        "last_login_at": row["last_login_at"] or "",
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }
    if include_admin:
        data["admin_note"] = row["admin_note"] or ""
    return data


def ssl_application_response(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "applicant_account": row["applicant_account"],
        "full_name": row["full_name"] if "full_name" in row.keys() else "",
        "grade": row["grade"] if "grade" in row.keys() else "",
        "self_intro": row["self_intro"],
        "interview_direction": row["interview_direction"],
        "interview_time": row["interview_time"],
        "status": row["status"],
        "interview_location": row["interview_location"] or "",
        "rejection_reason": row["rejection_reason"] or "",
        "reviewed_by": row["reviewed_by"] or "",
        "reviewed_at": row["reviewed_at"] or "",
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def site_message_response(row: sqlite3.Row) -> dict[str, Any]:
    read_at = row["read_at"] or ""
    return {
        "id": row["id"],
        "category": row["category"],
        "title": row["title"],
        "content": row["content"],
        "related_id": row["related_id"] or "",
        "read_at": read_at,
        "is_read": bool(read_at),
        "created_at": row["created_at"],
    }


def require_active_site_account(conn: sqlite3.Connection, account: str) -> sqlite3.Row:
    normalized_account = account.strip()
    if not normalized_account:
        raise HTTPException(status_code=401, detail="请先登录账号")
    row = conn.execute(
        "SELECT account, full_name, grade, is_disabled FROM site_account WHERE account = ?",
        (normalized_account,),
    ).fetchone()
    if row is None:
        raise HTTPException(status_code=401, detail="请先注册并登录 PRINTK 战队账号")
    if row["is_disabled"]:
        raise HTTPException(status_code=403, detail="账号已停用，请联系管理员")
    return row


def account_admin_log(conn: sqlite3.Connection, account: str, action: str, detail: str = "") -> None:
    timestamp = now_iso()
    conn.execute(
        """
        INSERT INTO site_account_admin_log (id, account, action, detail, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (uuid.uuid4().hex, account, action, detail, timestamp),
    )


def forum_author_name(row: sqlite3.Row) -> str:
    return row["full_name"] or row["author_account"]


def forum_post_response(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "title": row["title"],
        "content": row["content"],
        "author_account": row["author_account"],
        "author_name": forum_author_name(row),
        "status": row["status"] if "status" in row.keys() else FORUM_PUBLIC_STATUS,
        "reject_reason": row["reject_reason"] if "reject_reason" in row.keys() else "",
        "reviewed_by": row["reviewed_by"] if "reviewed_by" in row.keys() else "",
        "reviewed_at": row["reviewed_at"] if "reviewed_at" in row.keys() else "",
        "deleted_at": row["deleted_at"] if "deleted_at" in row.keys() else "",
        "is_pinned": bool(row["is_pinned"]) if "is_pinned" in row.keys() else False,
        "is_locked": bool(row["is_locked"]) if "is_locked" in row.keys() else False,
        "reply_count": row["reply_count"] if "reply_count" in row.keys() else 0,
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def forum_reply_response(row: sqlite3.Row) -> dict[str, str]:
    return {
        "id": row["id"],
        "post_id": row["post_id"],
        "content": row["content"],
        "author_account": row["author_account"],
        "author_name": forum_author_name(row),
        "status": row["status"] if "status" in row.keys() else FORUM_PUBLIC_STATUS,
        "reject_reason": row["reject_reason"] if "reject_reason" in row.keys() else "",
        "reviewed_by": row["reviewed_by"] if "reviewed_by" in row.keys() else "",
        "reviewed_at": row["reviewed_at"] if "reviewed_at" in row.keys() else "",
        "deleted_at": row["deleted_at"] if "deleted_at" in row.keys() else "",
        "created_at": row["created_at"],
        "updated_at": row["updated_at"] if "updated_at" in row.keys() and row["updated_at"] else row["created_at"],
    }


def normalize_forum_title(value: str) -> str:
    title = value.strip()
    if len(title) < 2 or len(title) > 80:
        raise HTTPException(status_code=400, detail="标题长度需为 2 到 80 个字符")
    return title


def normalize_forum_content(value: str, field_name: str = "内容", max_length: int = 5000) -> str:
    content = value.strip()
    if not content:
        raise HTTPException(status_code=400, detail=f"{field_name}不能为空")
    if len(content) > max_length:
        raise HTTPException(status_code=400, detail=f"{field_name}不能超过 {max_length} 个字符")
    return content


def ensure_forum_author(conn: sqlite3.Connection, account: str) -> str:
    normalized_account = normalize_account(account)
    row = conn.execute("SELECT account, is_disabled FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="账号不存在，请重新登录")
    if row["is_disabled"]:
        raise HTTPException(status_code=403, detail="账号已停用，请联系管理员")
    return normalized_account


def normalize_forum_status(value: str) -> str:
    status = value.strip()
    if status not in FORUM_CONTENT_STATUSES:
        raise HTTPException(status_code=400, detail="论坛状态格式错误")
    return status


def market_author_name(row: sqlite3.Row) -> str:
    return row["full_name"] or row["author_account"]


def market_status_label(status: str) -> str:
    return {
        "available": "可联系流转",
        "sold": "已出",
        "delisted": "已下架",
    }.get(status, "可联系流转")


def market_tags(value: str) -> list[str]:
    try:
        tags = json.loads(value or "[]")
    except json.JSONDecodeError:
        tags = []
    if not isinstance(tags, list):
        return []
    return [str(tag) for tag in tags if str(tag).strip()]


def flea_market_item_response(row: sqlite3.Row) -> dict[str, Any]:
    status = row["status"] if row["status"] in MARKET_ITEM_STATUSES else "available"
    return {
        "id": row["id"],
        "name": row["name"],
        "image_src": row["image_src"] or "/robots/engineering-robot.png",
        "image_alt": row["image_alt"] or row["name"],
        "author_account": row["author_account"],
        "owner": market_author_name(row),
        "team": row["team"] or row["department"] or "",
        "location": row["location"],
        "status": status,
        "status_text": market_status_label(status),
        "summary": row["summary"],
        "detail": row["detail"],
        "contact": row["contact"],
        "tags": market_tags(row["tags"]),
        "delisted_at": row["delisted_at"] if "delisted_at" in row.keys() else "",
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def normalize_market_name(value: str) -> str:
    name = value.strip()
    if len(name) < 2 or len(name) > 60:
        raise HTTPException(status_code=400, detail="物品名称长度需为 2 到 60 个字符")
    return name


def normalize_market_text(value: str, field_name: str, max_length: int) -> str:
    text = value.strip()
    if not text:
        raise HTTPException(status_code=400, detail=f"{field_name}不能为空")
    if len(text) > max_length:
        raise HTTPException(status_code=400, detail=f"{field_name}不能超过 {max_length} 个字符")
    return text


def normalize_market_image_src(value: str) -> str:
    image_src = value.strip() or "/robots/engineering-robot.png"
    if len(image_src) > 200:
        raise HTTPException(status_code=400, detail="图片路径不能超过 200 个字符")
    if not image_src.startswith("/"):
        raise HTTPException(status_code=400, detail="图片路径需使用站内路径")
    return image_src


def normalize_market_tags(value: str) -> str:
    raw_tags = value.replace("，", ",").replace("、", ",").split(",")
    tags: list[str] = []
    for raw_tag in raw_tags:
        tag = raw_tag.strip()
        if not tag:
            continue
        if len(tag) > 16:
            raise HTTPException(status_code=400, detail="单个标签不能超过 16 个字符")
        if tag not in tags:
            tags.append(tag)
    if len(tags) > 6:
        raise HTTPException(status_code=400, detail="标签最多填写 6 个")
    return json.dumps(tags, ensure_ascii=False)


def normalize_market_status(value: str) -> str:
    status = value.strip()
    if status not in MARKET_ITEM_STATUSES:
        raise HTTPException(status_code=400, detail="物品状态格式错误")
    return status


def ensure_market_author(conn: sqlite3.Connection, account: str) -> str:
    return ensure_forum_author(conn, account)


def validate_site_password(password: str) -> str:
    if len(password) < 6 or len(password) > 72:
        raise HTTPException(status_code=400, detail="密码长度需为 6 到 72 个字符")
    return password


def hash_site_password(password: str) -> str:
    salt = os.urandom(16).hex()
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 200_000).hex()
    return f"pbkdf2_sha256${salt}${digest}"


def verify_site_password(password: str, stored_hash: str) -> bool:
    try:
        algorithm, salt, digest = stored_hash.split("$", 2)
    except ValueError:
        return False
    if algorithm != "pbkdf2_sha256":
        return False
    current_digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 200_000).hex()
    return hmac.compare_digest(current_digest, digest)


def verify_token(token: str, allowed_roles: set[str]) -> str:
    try:
        decoded = base64.urlsafe_b64decode(token.encode("utf-8")).decode("utf-8")
        role, timestamp, signature = decoded.split(":", 2)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=401, detail="登录状态无效") from exc
    payload = f"{role}:{timestamp}"
    expected = hmac.new(SECRET_KEY.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(signature, expected):
        raise HTTPException(status_code=401, detail="登录状态无效")
    if role not in allowed_roles:
        raise HTTPException(status_code=403, detail="权限不足")
    return role


def require_admin(authorization: str | None = Header(default=None)) -> str:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="请先登录管理员后台")
    return verify_token(authorization.removeprefix("Bearer ").strip(), {"admin"})


def require_site_plan_editor(payload: "SeasonPlanRequest") -> str:
    account = normalize_account(payload.editor_account)
    with db_connection() as conn:
        row = conn.execute(
            "SELECT account, permission_level, is_disabled FROM site_account WHERE account = ?",
            (account,),
        ).fetchone()
    if row is None:
        raise HTTPException(status_code=401, detail="请先登录账号")
    if row["is_disabled"]:
        raise HTTPException(status_code=403, detail="账号已停用，请联系管理员")
    if row["permission_level"] not in PLAN_EDITOR_PERMISSION_OPTIONS:
        raise HTTPException(status_code=403, detail="当前账号没有赛季规划编辑权限")
    return account


def log_process(stage: str, level: str, message: str, batch_id: str | None = None) -> None:
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO process_log (id, batch_id, stage, level, message, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (uuid.uuid4().hex, batch_id, stage, level, message, now_iso()),
        )


def stage_path(stage: str) -> Path:
    mapping = {
        "unregistered": UNREGISTERED_DIR,
        "pending_review": PENDING_REVIEW_DIR,
        "in_stock": IN_STOCK_DIR,
        "out_stock": OUT_STOCK_DIR,
        "processing": PROCESSING_DIR,
        "duplicate_invoice": DUPLICATE_ARCHIVE_DIR,
        "parse_failed": PARSE_FAILED_DIR,
        "validation_failed": VALIDATION_FAILED_DIR,
        "review_rejected": REVIEW_REJECTED_DIR,
    }
    return mapping[stage]


def batch_form_path(batch_id: str, folder_stage: str) -> Path:
    return stage_path(folder_stage) / batch_id / "form.xlsx"


def move_batch_folder(batch_id: str, from_stage: str, to_stage: str) -> None:
    source = stage_path(from_stage) / batch_id
    target = stage_path(to_stage) / batch_id
    if not source.exists():
        return
    target.parent.mkdir(parents=True, exist_ok=True)
    if target.exists():
        shutil.rmtree(target)
    shutil.move(str(source), str(target))


def write_meta(batch_dir: Path, meta: dict[str, Any]) -> None:
    (batch_dir / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_excel(form_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(form_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("表格为空")
    headers = [normalize_header(cell) for cell in rows[0]]
    for required_header in REQUIRED_HEADERS:
        if normalize_header(required_header) not in headers:
            raise ValueError(f"缺少必填列：{required_header}")
    header_indexes = {
        name: headers.index(normalize_header(name))
        for name in HEADER_MAP
        if normalize_header(name) in headers
    }

    parsed_rows: list[dict[str, Any]] = []
    for row_no, row in enumerate(rows[1:], start=2):
        row_values = list(row)
        if not any(value not in (None, "") for value in row_values):
            continue

        def cell(header_name: str) -> Any:
            index = header_indexes.get(header_name)
            if index is None or index >= len(row_values):
                return None
            return row_values[index]

        record = {
            "row_no": row_no,
            "purchase_date": normalize_date(cell("采购日期"), "采购日期", True),
            "item_name": normalize_text(cell("物资名称")),
            "spec_model": normalize_text(cell("规格型号")),
            "unit": normalize_text(cell("单位")),
            "quantity": normalize_number(cell("数量"), "数量", True),
            "unit_price": normalize_number(cell("单价"), "单价", True),
            "amount": normalize_number(cell("金额"), "金额", True),
            "supplier_name": normalize_text(cell("供应商")),
            "invoice_code": normalize_text(cell("发票代码")),
            "invoice_number": normalize_text(cell("发票号码")),
            "invoice_date": normalize_date(cell("发票日期"), "发票日期", False),
            "invoice_amount": normalize_number(cell("发票金额"), "发票金额", False),
            "remark": normalize_text(cell("备注")),
        }
        if not record["item_name"]:
            raise ValueError(f"第 {row_no} 行物资名称不能为空")
        if not record["unit"]:
            raise ValueError(f"第 {row_no} 行单位不能为空")
        if not record["invoice_number"]:
            raise ValueError(f"第 {row_no} 行发票号码不能为空")
        parsed_rows.append(record)

    if not parsed_rows:
        raise ValueError("表格没有有效数据")
    return parsed_rows


def invoice_exists(conn: sqlite3.Connection, invoice_number: str) -> tuple[bool, str]:
    pending = conn.execute(
        """
        SELECT batch_id FROM staged_purchase_record
        WHERE invoice_number = ? AND review_status = 'pending_review'
        LIMIT 1
        """,
        (invoice_number,),
    ).fetchone()
    if pending:
        return True, f"待入库阶段已存在相同发票号码，批次 {pending['batch_id']}"
    registry = conn.execute("SELECT batch_id FROM invoice_registry WHERE invoice_number = ? LIMIT 1", (invoice_number,)).fetchone()
    if registry:
        return True, f"库内已存在相同发票号码，批次 {registry['batch_id']}"
    reimbursement = conn.execute(
        "SELECT reimbursement_batch_id FROM reimbursement_record WHERE invoice_number = ? LIMIT 1",
        (invoice_number,),
    ).fetchone()
    if reimbursement:
        return True, f"已存在出库报销记录，出库批次 {reimbursement['reimbursement_batch_id']}"
    return False, ""


def fail_batch(batch_id: str, from_stage: str, target_stage: str, status: str, reason: str) -> None:
    move_batch_folder(batch_id, from_stage, target_stage)
    with db_connection() as conn:
        conn.execute(
            """
            UPDATE upload_batch
            SET status = ?, folder_stage = ?, form_file_path = ?, review_note = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, target_stage, str(batch_form_path(batch_id, target_stage)), reason, now_iso(), batch_id),
        )
    log_process("local_review", "error", reason, batch_id)


def process_batch(batch_row: sqlite3.Row) -> None:
    batch_id = batch_row["id"]
    move_batch_folder(batch_id, "unregistered", "processing")
    with db_connection() as conn:
        conn.execute(
            """
            UPDATE upload_batch
            SET status = 'processing', folder_stage = 'processing', form_file_path = ?, updated_at = ?
            WHERE id = ?
            """,
            (str(batch_form_path(batch_id, "processing")), now_iso(), batch_id),
        )
    form_path = PROCESSING_DIR / batch_id / "form.xlsx"
    try:
        parsed_rows = parse_excel(form_path)
    except ValueError as exc:
        fail_batch(batch_id, "processing", "validation_failed", "validation_failed", str(exc))
        return
    except Exception as exc:  # noqa: BLE001
        fail_batch(batch_id, "processing", "parse_failed", "parse_failed", f"解析失败：{exc}")
        return

    invoice_numbers = [row["invoice_number"] for row in parsed_rows]
    if len(invoice_numbers) != len(set(invoice_numbers)):
        fail_batch(batch_id, "processing", "duplicate_invoice", "duplicate_invoice", "当前表格存在重复发票号码")
        return

    with db_connection() as conn:
        for row in parsed_rows:
            exists, reason = invoice_exists(conn, row["invoice_number"])
            if exists:
                fail_batch(batch_id, "processing", "duplicate_invoice", "duplicate_invoice", reason)
                return
        current_time = now_iso()
        for row in parsed_rows:
            conn.execute(
                """
                INSERT INTO staged_purchase_record (
                    id, batch_id, row_no, purchase_date, item_name, spec_model, unit,
                    quantity, unit_price, amount, supplier_name, invoice_code,
                    invoice_number, invoice_date, invoice_amount, remark,
                    review_status, review_note, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    uuid.uuid4().hex,
                    batch_id,
                    row["row_no"],
                    row["purchase_date"],
                    row["item_name"],
                    row["spec_model"],
                    row["unit"],
                    row["quantity"],
                    row["unit_price"],
                    row["amount"],
                    row["supplier_name"],
                    row["invoice_code"],
                    row["invoice_number"],
                    row["invoice_date"],
                    row["invoice_amount"],
                    row["remark"],
                    "pending_review",
                    "",
                    current_time,
                    current_time,
                ),
            )
        conn.execute(
            """
            UPDATE upload_batch
            SET status = 'pending_review', folder_stage = 'pending_review', form_file_path = ?, review_note = '', updated_at = ?
            WHERE id = ?
            """,
            (str(batch_form_path(batch_id, "pending_review")), current_time, batch_id),
        )
    move_batch_folder(batch_id, "processing", "pending_review")
    log_process("local_review", "info", "批次审核通过，已进入待入库队列", batch_id)


def run_agent_once() -> int:
    if not AGENT_LOCK.acquire(blocking=False):
        return 0
    try:
        with db_connection() as conn:
            batches = conn.execute(
                """
                SELECT * FROM upload_batch
                WHERE status = 'unregistered' AND folder_stage = 'unregistered'
                ORDER BY submitted_at ASC
                """
            ).fetchall()
        for batch in batches:
            process_batch(batch)
        return len(batches)
    finally:
        AGENT_LOCK.release()


def confirm_staged_rows(batch_id: str, row_ids: list[str]) -> int:
    if not row_ids:
        return 0
    with db_connection() as conn:
        placeholders = ",".join("?" for _ in row_ids)
        rows = conn.execute(
            f"""
            SELECT * FROM staged_purchase_record
            WHERE batch_id = ? AND id IN ({placeholders}) AND review_status = 'pending_review'
            """,
            [batch_id, *row_ids],
        ).fetchall()
        current_time = now_iso()
        for row in rows:
            purchase_record_id = uuid.uuid4().hex
            conn.execute(
                """
                INSERT INTO purchase_record (
                    id, batch_id, staged_record_id, row_no, purchase_date, item_name, spec_model,
                    unit, quantity, unit_price, amount, supplier_name, invoice_code,
                    invoice_number, invoice_date, invoice_amount, remark,
                    stock_status, reimbursed_at, reimbursement_batch_id, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    purchase_record_id,
                    row["batch_id"],
                    row["id"],
                    row["row_no"],
                    row["purchase_date"],
                    row["item_name"],
                    row["spec_model"],
                    row["unit"],
                    row["quantity"],
                    row["unit_price"],
                    row["amount"],
                    row["supplier_name"],
                    row["invoice_code"],
                    row["invoice_number"],
                    row["invoice_date"],
                    row["invoice_amount"],
                    row["remark"],
                    "in_stock",
                    "",
                    "",
                    current_time,
                    current_time,
                ),
            )
            conn.execute(
                """
                INSERT OR IGNORE INTO invoice_registry (id, invoice_number, batch_id, purchase_record_id, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (uuid.uuid4().hex, row["invoice_number"], batch_id, purchase_record_id, current_time),
            )
            conn.execute(
                """
                UPDATE staged_purchase_record
                SET review_status = 'confirmed', review_note = '管理员确认入库', updated_at = ?
                WHERE id = ?
                """,
                (current_time, row["id"]),
            )
    export_master_sheets()
    return len(rows)


def reject_staged_rows(batch_id: str, row_ids: list[str], note: str) -> int:
    if not row_ids:
        return 0
    with db_connection() as conn:
        placeholders = ",".join("?" for _ in row_ids)
        result = conn.execute(
            f"""
            UPDATE staged_purchase_record
            SET review_status = 'rejected', review_note = ?, updated_at = ?
            WHERE batch_id = ? AND id IN ({placeholders}) AND review_status = 'pending_review'
            """,
            [note, now_iso(), batch_id, *row_ids],
        )
    return result.rowcount


def finalize_batch_review(batch_id: str) -> tuple[bool, str]:
    with db_connection() as conn:
        batch = conn.execute("SELECT * FROM upload_batch WHERE id = ?", (batch_id,)).fetchone()
        if batch is None:
            return False, "批次不存在"
        counts = conn.execute(
            """
            SELECT
                SUM(CASE WHEN review_status = 'pending_review' THEN 1 ELSE 0 END) AS pending_count,
                SUM(CASE WHEN review_status = 'confirmed' THEN 1 ELSE 0 END) AS confirmed_count,
                SUM(CASE WHEN review_status = 'rejected' THEN 1 ELSE 0 END) AS rejected_count
            FROM staged_purchase_record
            WHERE batch_id = ?
            """,
            (batch_id,),
        ).fetchone()
        if counts["pending_count"]:
            return False, "仍有待确认明细"
        current_time = now_iso()
        if counts["confirmed_count"]:
            move_batch_folder(batch_id, "pending_review", "in_stock")
            conn.execute(
                """
                UPDATE upload_batch
                SET status = 'in_stock', folder_stage = 'in_stock', form_file_path = ?, updated_at = ?
                WHERE id = ?
                """,
                (str(batch_form_path(batch_id, "in_stock")), current_time, batch_id),
            )
            export_master_sheets()
            return True, "批次已进入库内"
        move_batch_folder(batch_id, "pending_review", "review_rejected")
        conn.execute(
            """
            UPDATE upload_batch
            SET status = 'review_rejected', folder_stage = 'review_rejected', form_file_path = ?, updated_at = ?
            WHERE id = ?
            """,
            (str(batch_form_path(batch_id, "review_rejected")), current_time, batch_id),
        )
    return True, "批次已全部打回"


def export_records_to_excel(path: Path, rows: list[sqlite3.Row]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据"
    sheet.append(
        [
            "批次编号",
            "表格行号",
            "采购日期",
            "物资名称",
            "规格型号",
            "单位",
            "数量",
            "单价",
            "金额",
            "供应商",
            "发票代码",
            "发票号码",
            "发票日期",
            "发票金额",
            "备注",
            "库存状态",
            "出库时间",
        ]
    )
    for row in rows:
        sheet.append(
            [
                row["batch_id"],
                row["row_no"],
                row["purchase_date"],
                row["item_name"],
                row["spec_model"],
                row["unit"],
                row["quantity"],
                row["unit_price"],
                row["amount"],
                row["supplier_name"],
                row["invoice_code"],
                row["invoice_number"],
                row["invoice_date"],
                row["invoice_amount"],
                row["remark"],
                row["stock_status"],
                row["reimbursed_at"],
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def export_master_sheets() -> None:
    with db_connection() as conn:
        in_stock_rows = conn.execute(
            """
            SELECT pr.* FROM purchase_record pr
            JOIN upload_batch ub ON ub.id = pr.batch_id
            WHERE pr.stock_status = 'in_stock' AND ub.folder_stage = 'in_stock'
            ORDER BY created_at ASC
            """
        ).fetchall()
        out_stock_rows = conn.execute(
            """
            SELECT pr.* FROM purchase_record pr
            JOIN upload_batch ub ON ub.id = pr.batch_id
            WHERE pr.stock_status = 'out_stock' AND ub.folder_stage = 'out_stock'
            ORDER BY updated_at ASC
            """
        ).fetchall()
    export_records_to_excel(IN_STOCK_MASTER_DIR / "库内总表.xlsx", in_stock_rows)
    export_records_to_excel(OUT_STOCK_MASTER_DIR / "出库历史总表.xlsx", out_stock_rows)


def build_template_workbook() -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "发票模板"
    sheet.append(list(HEADER_MAP.keys()))
    sheet.append(["2026-06-18", "A4纸", "80g", "包", 2, 25, 50, "文具店", "FP01", "INV-0001", "2026-06-18", 50, "示例数据"])
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def export_reimbursement(record_ids: list[str]) -> tuple[bool, str]:
    if not record_ids:
        return False, "请选择需要出库的库内明细"
    with db_connection() as conn:
        placeholders = ",".join("?" for _ in record_ids)
        rows = conn.execute(
            f"""
            SELECT pr.* FROM purchase_record pr
            JOIN upload_batch ub ON ub.id = pr.batch_id
            WHERE pr.id IN ({placeholders}) AND pr.stock_status = 'in_stock' AND ub.folder_stage = 'in_stock'
            ORDER BY batch_id ASC, row_no ASC
            """,
            record_ids,
        ).fetchall()
        if not rows:
            return False, "没有可出库的库内明细"
        reimbursement_id = f"RB{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:4]}"
        export_path = REIMBURSEMENT_EXPORT_DIR / f"{reimbursement_id}.xlsx"
        export_records_to_excel(export_path, rows)
        total_amount = sum(row["amount"] for row in rows)
        current_time = now_iso()
        conn.execute(
            """
            INSERT INTO reimbursement_batch (
                id, batch_name, extracted_by, extracted_at, record_count,
                total_amount, export_file_path, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (reimbursement_id, f"报销提取_{reimbursement_id}", "管理员", current_time, len(rows), total_amount, str(export_path), current_time),
        )
        affected_batches: set[str] = set()
        for row in rows:
            conn.execute(
                """
                INSERT INTO reimbursement_record (
                    id, reimbursement_batch_id, purchase_record_id,
                    invoice_number, reimbursed_at, created_at
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (uuid.uuid4().hex, reimbursement_id, row["id"], row["invoice_number"], current_time, current_time),
            )
            conn.execute(
                """
                UPDATE purchase_record
                SET stock_status = 'out_stock', reimbursed_at = ?, reimbursement_batch_id = ?, updated_at = ?
                WHERE id = ?
                """,
                (current_time, reimbursement_id, current_time, row["id"]),
            )
            affected_batches.add(row["batch_id"])
        for batch_id in affected_batches:
            remaining = conn.execute(
                "SELECT COUNT(*) AS total FROM purchase_record WHERE batch_id = ? AND stock_status = 'in_stock'",
                (batch_id,),
            ).fetchone()["total"]
            if remaining == 0:
                move_batch_folder(batch_id, "in_stock", "out_stock")
                conn.execute(
                    """
                    UPDATE upload_batch
                    SET status = 'out_stock', folder_stage = 'out_stock', form_file_path = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (str(batch_form_path(batch_id, "out_stock")), current_time, batch_id),
                )
    export_master_sheets()
    return True, reimbursement_id


def dashboard_data() -> dict[str, Any]:
    with db_connection() as conn:
        counts = {
            "unregistered": conn.execute("SELECT COUNT(*) AS total FROM upload_batch WHERE status = 'unregistered'").fetchone()["total"],
            "pending_review": conn.execute("SELECT COUNT(*) AS total FROM upload_batch WHERE folder_stage = 'pending_review'").fetchone()["total"],
            "in_stock": conn.execute(
                """
                SELECT COUNT(*) AS total FROM purchase_record pr
                JOIN upload_batch ub ON ub.id = pr.batch_id
                WHERE pr.stock_status = 'in_stock' AND ub.folder_stage = 'in_stock'
                """
            ).fetchone()["total"],
            "out_stock": conn.execute(
                """
                SELECT COUNT(*) AS total FROM purchase_record pr
                JOIN upload_batch ub ON ub.id = pr.batch_id
                WHERE pr.stock_status = 'out_stock' AND ub.folder_stage = 'out_stock'
                """
            ).fetchone()["total"],
        }
        pending_batches = [row_to_dict(row) for row in conn.execute(
            """
            SELECT
                ub.*,
                SUM(CASE WHEN spr.review_status = 'pending_review' THEN 1 ELSE 0 END) AS pending_rows,
                SUM(CASE WHEN spr.review_status = 'confirmed' THEN 1 ELSE 0 END) AS confirmed_rows,
                SUM(CASE WHEN spr.review_status = 'rejected' THEN 1 ELSE 0 END) AS rejected_rows
            FROM upload_batch ub
            LEFT JOIN staged_purchase_record spr ON spr.batch_id = ub.id
            WHERE ub.folder_stage = 'pending_review'
            GROUP BY ub.id
            ORDER BY ub.submitted_at ASC
            """
        ).fetchall()]
        in_stock_rows = [row_to_dict(row) for row in conn.execute(
            """
            SELECT pr.* FROM purchase_record pr
            JOIN upload_batch ub ON ub.id = pr.batch_id
            WHERE pr.stock_status = 'in_stock' AND ub.folder_stage = 'in_stock'
            ORDER BY pr.created_at ASC
            """
        ).fetchall()]
        reimbursement_batches = [row_to_dict(row) for row in conn.execute(
            "SELECT * FROM reimbursement_batch ORDER BY created_at DESC LIMIT 10"
        ).fetchall()]
        logs = [row_to_dict(row) for row in conn.execute("SELECT * FROM process_log ORDER BY created_at DESC LIMIT 20").fetchall()]
    return {"counts": counts, "pending_batches": pending_batches, "in_stock_rows": in_stock_rows, "reimbursement_batches": reimbursement_batches, "logs": logs}


def validate_plan_period(season_year: int, month: int) -> None:
    if season_year < 2020 or season_year > 2100:
        raise HTTPException(status_code=400, detail="赛季年份格式错误")
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="月份格式错误")


def list_season_plan(season_year: int, month: int) -> list[dict[str, Any]]:
    validate_plan_period(season_year, month)
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                id, season_year, month, robot_type, task_title, status, target,
                assignee_account, is_completed, display_order, updated_at
            FROM season_plan
            WHERE season_year = ? AND month = ?
                AND robot_type IN ('英雄兵种', '步兵兵种', '工程兵种', '哨兵兵种')
            ORDER BY display_order ASC, robot_type ASC, task_title ASC
            """,
            (season_year, month),
        ).fetchall()
    plans = []
    for row in rows:
        plan = row_to_dict(row)
        plan["is_completed"] = bool(plan["is_completed"])
        plans.append(plan)
    return plans


def save_season_plan(season_year: int, month: int, plans: list["SeasonPlanItem"]) -> list[dict[str, Any]]:
    validate_plan_period(season_year, month)
    if not plans:
        raise HTTPException(status_code=400, detail="计划不能为空")
    timestamp = now_iso()
    with db_connection() as conn:
        conn.execute("DELETE FROM season_plan WHERE season_year = ? AND month = ?", (season_year, month))
        for index, plan in enumerate(plans, start=1):
            robot_type = plan.robot_type.strip()
            task_title = plan.task_title.strip()
            status = plan.status.strip()
            target = plan.target.strip()
            assignee_account = plan.assignee_account.strip()
            if robot_type not in SEASON_PLAN_ROBOT_TYPES:
                raise HTTPException(status_code=400, detail="兵种格式错误")
            if not task_title or not status or not target:
                raise HTTPException(status_code=400, detail="兵种、任务、状态、目标不能为空")
            if assignee_account:
                assignee = conn.execute(
                    "SELECT account, is_disabled FROM site_account WHERE account = ?",
                    (assignee_account,),
                ).fetchone()
                if assignee is None:
                    raise HTTPException(status_code=400, detail=f"执行人账号不存在：{assignee_account}")
                if assignee["is_disabled"]:
                    raise HTTPException(status_code=400, detail=f"执行人账号已停用：{assignee_account}")
            conn.execute(
                """
                INSERT INTO season_plan (
                    id, season_year, month, group_name, robot_type, task_title, status, target,
                    assignee_account, is_completed, display_order, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    uuid.uuid4().hex,
                    season_year,
                    month,
                    f"{robot_type}：{task_title}",
                    robot_type,
                    task_title,
                    status,
                    target,
                    assignee_account,
                    1 if plan.is_completed else 0,
                    index,
                    timestamp,
                    timestamp,
                ),
            )
    return list_season_plan(season_year, month)


def homepage_asset_response(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "kind": row["kind"],
        "url": row["url"],
        "original_filename": row["original_filename"] or "",
        "mime_type": row["mime_type"] or "",
        "size_bytes": row["size_bytes"] or 0,
        "alt": row["alt"] or "",
        "display_order": row["display_order"] or 0,
        "is_enabled": bool(row["is_enabled"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def homepage_quote_response(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "text": row["text"],
        "source": row["source"] or "",
        "display_order": row["display_order"] or 0,
        "is_enabled": bool(row["is_enabled"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def recruitment_faq_response(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "question": row["question"],
        "answer": row["answer"],
        "display_order": row["display_order"],
        "is_enabled": bool(row["is_enabled"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def recruitment_question_response(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "author_account": row["author_account"],
        "author_name": row["author_name"] or row["author_account"],
        "content": row["content"],
        "created_at": row["created_at"],
    }


def homepage_recruitment_banner_response(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "text": row["text"],
        "action_text": row["action_text"],
        "is_enabled": bool(row["is_enabled"]),
        "updated_at": row["updated_at"],
    }


def ordered_homepage_awards(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    awards = [
        {
            **award,
            "display_order": award.get("display_order", index + 1),
        }
        for index, award in enumerate(value)
        if isinstance(award, dict)
    ]
    awards.sort(key=lambda award: award["display_order"])
    for index, award in enumerate(awards):
        award["display_order"] = index + 1
    return awards


def homepage_profile_response(row: sqlite3.Row) -> dict[str, Any]:
    try:
        stats = json.loads(row["stats_json"] or "[]")
    except json.JSONDecodeError:
        stats = HOME_PROFILE_STATS
    try:
        awards = ordered_homepage_awards(json.loads(row["awards_json"] or "[]"))
    except json.JSONDecodeError:
        awards = HOME_PROFILE_AWARDS
    saved_recruitment: dict[str, Any] = {}
    try:
        saved_recruitment = json.loads(row["recruitment_json"] or "{}")
        recruitment = {**HOME_RECRUITMENT_CONTENT, **saved_recruitment} if isinstance(saved_recruitment, dict) else HOME_RECRUITMENT_CONTENT
    except json.JSONDecodeError:
        recruitment = HOME_RECRUITMENT_CONTENT
    saved_events = saved_recruitment.get("events") if isinstance(saved_recruitment, dict) else None
    if not isinstance(saved_events, list) or len(saved_events) != 3:
        recruitment["events"] = [
            {
                **HOME_RECRUITMENT_CONTENT["events"][0],
                "kicker": recruitment["event_kicker"],
                "title": recruitment["event_title"],
                "description": recruitment["event_description"],
            },
            *[dict(event) for event in HOME_RECRUITMENT_CONTENT["events"][1:]],
        ]
    return {
        "team_name": row["team_name"],
        "team_intro": row["team_intro"],
        "stats": stats,
        "awards": awards,
        "recruitment": recruitment,
        "updated_at": row["updated_at"],
    }


def homepage_danmaku_response(row: sqlite3.Row) -> dict[str, Any]:
    image_key = row["image_key"] if "image_key" in row.keys() else row["image_src"]
    author_name = row["author_name"] if "author_name" in row.keys() else ""
    author_account = row["author_account"] if "author_account" in row.keys() else ""
    account_full_name = row["account_full_name"] if "account_full_name" in row.keys() else ""
    return {
        "id": row["id"],
        "imageKey": image_key,
        "imageSrc": row["image_src"],
        "authorAccount": author_account,
        "authorName": author_name or account_full_name or author_account,
        "text": row["text"],
        "track": row["track"],
        "color": row["color"],
        "status": row["status"] if "status" in row.keys() else "approved",
        "reviewedBy": row["reviewed_by"] if "reviewed_by" in row.keys() else "",
        "reviewedAt": row["reviewed_at"] if "reviewed_at" in row.keys() else "",
        "createdAt": row["created_at_ms"],
        "created_at": row["created_at"],
        "duration": row["duration"],
        "delay": row["delay"],
    }


def get_homepage_content(include_disabled: bool = False) -> dict[str, Any]:
    enabled_clause = "" if include_disabled else "AND is_enabled = 1"
    with db_connection() as conn:
        videos = conn.execute(
            f"""
            SELECT *
            FROM homepage_asset
            WHERE kind = 'video' {enabled_clause}
            ORDER BY display_order ASC, created_at DESC
            """
        ).fetchall()
        images = conn.execute(
            f"""
            SELECT *
            FROM homepage_asset
            WHERE kind = 'image' {enabled_clause}
            ORDER BY display_order ASC, created_at ASC
            """
        ).fetchall()
        quotes = conn.execute(
            f"""
            SELECT *
            FROM homepage_quote
            WHERE 1 = 1 {enabled_clause}
            ORDER BY display_order ASC, created_at ASC
            """
        ).fetchall()
        recruitment_banner = conn.execute(
            """
            SELECT * FROM homepage_recruitment_banner
            WHERE id = 'recruitment'
            """
        ).fetchone()
        campus_banner = conn.execute(
            """
            SELECT * FROM homepage_recruitment_banner
            WHERE id = 'campus-competition'
            """
        ).fetchone()
        profile = conn.execute("SELECT * FROM homepage_profile WHERE id = 'profile'").fetchone()
        faqs = conn.execute(
            f"SELECT * FROM recruitment_faq WHERE 1 = 1 {enabled_clause} ORDER BY display_order ASC, created_at ASC"
        ).fetchall()
        questions = conn.execute(
            """
            SELECT question.*, account.full_name AS author_name
            FROM recruitment_question AS question
            LEFT JOIN site_account AS account ON account.account = question.author_account
            ORDER BY question.created_at DESC
            """
        ).fetchall() if include_disabled else []
    video_items = [homepage_asset_response(row) for row in videos]
    return {
        "video": video_items[0] if video_items else None,
        "videos": video_items,
        "images": [homepage_asset_response(row) for row in images],
        "quotes": [homepage_quote_response(row) for row in quotes],
        "recruitment_banner": homepage_recruitment_banner_response(recruitment_banner)
        if recruitment_banner and (include_disabled or recruitment_banner["is_enabled"])
        else None,
        "campus_banner": homepage_recruitment_banner_response(campus_banner)
        if campus_banner and (include_disabled or campus_banner["is_enabled"])
        else None,
        "profile": homepage_profile_response(profile),
        "faqs": [recruitment_faq_response(row) for row in faqs],
        "recruitment_questions": [recruitment_question_response(row) for row in questions],
    }


def create_recruitment_question(payload: "RecruitmentQuestionCreate") -> dict[str, Any]:
    account = payload.author_account.strip()
    content = normalize_limited_text(payload.content, "问题", 500)
    if not content:
        raise HTTPException(status_code=400, detail="请输入招新问题")
    timestamp = now_iso()
    question_id = uuid.uuid4().hex
    with db_connection() as conn:
        require_active_site_account(conn, account)
        conn.execute(
            "INSERT INTO recruitment_question (id, author_account, content, created_at) VALUES (?, ?, ?, ?)",
            (question_id, account, content, timestamp),
        )
        row = conn.execute(
            """
            SELECT question.*, account.full_name AS author_name
            FROM recruitment_question AS question
            LEFT JOIN site_account AS account ON account.account = question.author_account
            WHERE question.id = ?
            """,
            (question_id,),
        ).fetchone()
    return recruitment_question_response(row)


def save_recruitment_faq(payload: "RecruitmentFaqSave", faq_id: str | None = None) -> dict[str, Any]:
    question = normalize_limited_text(payload.question, "问题", 200)
    answer = normalize_limited_text(payload.answer, "回答", 1000)
    if not question or not answer:
        raise HTTPException(status_code=400, detail="问题和回答不能为空")
    item_id = faq_id or uuid.uuid4().hex
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute("SELECT id FROM recruitment_faq WHERE id = ?", (item_id,)).fetchone()
        if faq_id and not existing:
            raise HTTPException(status_code=404, detail="QA 不存在")
        if existing:
            conn.execute(
                "UPDATE recruitment_faq SET question = ?, answer = ?, display_order = ?, is_enabled = ?, updated_at = ? WHERE id = ?",
                (question, answer, payload.display_order, int(payload.is_enabled), timestamp, item_id),
            )
        else:
            conn.execute(
                "INSERT INTO recruitment_faq (id, question, answer, display_order, is_enabled, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (item_id, question, answer, payload.display_order, int(payload.is_enabled), timestamp, timestamp),
            )
        row = conn.execute("SELECT * FROM recruitment_faq WHERE id = ?", (item_id,)).fetchone()
    return recruitment_faq_response(row)


def delete_recruitment_faq(faq_id: str) -> dict[str, str]:
    with db_connection() as conn:
        if not conn.execute("SELECT id FROM recruitment_faq WHERE id = ?", (faq_id,)).fetchone():
            raise HTTPException(status_code=404, detail="QA 不存在")
        conn.execute("DELETE FROM recruitment_faq WHERE id = ?", (faq_id,))
    return {"id": faq_id}


def delete_recruitment_question(question_id: str) -> dict[str, str]:
    with db_connection() as conn:
        if not conn.execute("SELECT id FROM recruitment_question WHERE id = ?", (question_id,)).fetchone():
            raise HTTPException(status_code=404, detail="提问不存在")
        conn.execute("DELETE FROM recruitment_question WHERE id = ?", (question_id,))
    return {"id": question_id}


def update_homepage_recruitment_banner(payload: "HomepageRecruitmentBannerUpdate") -> dict[str, Any]:
    text = normalize_limited_text(payload.text, "招新公告文案", 120)
    action_text = normalize_limited_text(payload.action_text, "栏尾文案", 32)
    if not text:
        raise HTTPException(status_code=400, detail="招新公告文案不能为空")
    if not action_text:
        raise HTTPException(status_code=400, detail="栏尾文案不能为空")
    timestamp = now_iso()
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO homepage_recruitment_banner (id, text, action_text, is_enabled, created_at, updated_at)
            VALUES ('recruitment', ?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                text = excluded.text,
                action_text = excluded.action_text,
                is_enabled = excluded.is_enabled,
                updated_at = excluded.updated_at
            """,
            (text, action_text, 1 if payload.is_enabled else 0, timestamp, timestamp),
        )
        row = conn.execute("SELECT * FROM homepage_recruitment_banner WHERE id = 'recruitment'").fetchone()
    return homepage_recruitment_banner_response(row)


def update_homepage_campus_banner(payload: "HomepageRecruitmentBannerUpdate") -> dict[str, Any]:
    text = normalize_limited_text(payload.text, "校内赛公告文案", 120)
    action_text = normalize_limited_text(payload.action_text, "栏尾文案", 32)
    if not text:
        raise HTTPException(status_code=400, detail="校内赛公告文案不能为空")
    if not action_text:
        raise HTTPException(status_code=400, detail="栏尾文案不能为空")
    timestamp = now_iso()
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO homepage_recruitment_banner (id, text, action_text, is_enabled, created_at, updated_at)
            VALUES ('campus-competition', ?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                text = excluded.text,
                action_text = excluded.action_text,
                is_enabled = excluded.is_enabled,
                updated_at = excluded.updated_at
            """,
            (text, action_text, 1 if payload.is_enabled else 0, timestamp, timestamp),
        )
        row = conn.execute("SELECT * FROM homepage_recruitment_banner WHERE id = 'campus-competition'").fetchone()
    return homepage_recruitment_banner_response(row)


def update_homepage_profile(payload: "HomepageProfileUpdate") -> dict[str, Any]:
    team_name = normalize_limited_text(payload.team_name, "战队名", 80)
    team_intro = normalize_limited_text(payload.team_intro, "战队简介", 1000)
    if not team_name or not team_intro:
        raise HTTPException(status_code=400, detail="战队名和战队简介不能为空")
    if len(payload.stats) != 3:
        raise HTTPException(status_code=400, detail="首页概览需保留三个小窗口")
    if len(payload.awards) > 12:
        raise HTTPException(status_code=400, detail="奖项与荣誉最多配置 12 项")

    stats = []
    for item in payload.stats:
        value = normalize_limited_text(item.value, "概览数值", 24)
        label = normalize_limited_text(item.label, "概览名称", 32)
        if not value or not label:
            raise HTTPException(status_code=400, detail="概览数值和名称不能为空")
        stats.append({"value": value, "label": label})

    awards = []
    for item in payload.awards:
        title = normalize_limited_text(item.title, "奖项名称", 100)
        meta = normalize_limited_text(item.meta, "奖项说明", 240)
        image_url = normalize_limited_text(item.image_url, "奖项图片地址", 500)
        image_alt = normalize_limited_text(item.image_alt, "奖项图片说明", 160)
        display_order = normalize_homepage_order(item.display_order)
        if not title:
            continue
        awards.append({
            "title": title,
            "meta": meta,
            "image_url": image_url,
            "image_alt": image_alt,
            "display_order": display_order,
        })
    awards = ordered_homepage_awards(awards)

    recruitment = {
        "season_label": normalize_limited_text(payload.recruitment.season_label, "招新赛季标签", 40),
        "title": normalize_limited_text(payload.recruitment.title, "招新主标题", 120),
        "intro": normalize_limited_text(payload.recruitment.intro, "招新引导文案", 500),
        "event_kicker": normalize_limited_text(payload.recruitment.event_kicker, "赛事栏目标签", 60),
        "event_title": normalize_limited_text(payload.recruitment.event_title, "赛事介绍标题", 100),
        "event_description": normalize_limited_text(payload.recruitment.event_description, "赛事介绍", 1000),
        "events": [],
        "groups_kicker": normalize_limited_text(payload.recruitment.groups_kicker, "组别栏目标签", 60),
        "groups_title": normalize_limited_text(payload.recruitment.groups_title, "组别标题", 100),
        "groups": [],
        "qr_text": normalize_limited_text(payload.recruitment.qr_text, "二维码提示", 100),
    }
    if len(payload.recruitment.events) not in (0, 3):
        raise HTTPException(status_code=400, detail="赛事介绍需要配置 3 项")
    source_events = payload.recruitment.events or [
        HomepageRecruitmentEventItem(id="robomaster", name="RoboMaster", kicker=recruitment["event_kicker"], title=recruitment["event_title"], description=recruitment["event_description"]),
        HomepageRecruitmentEventItem(**HOME_RECRUITMENT_CONTENT["events"][1]),
        HomepageRecruitmentEventItem(**HOME_RECRUITMENT_CONTENT["events"][2]),
    ]
    for item in source_events:
        recruitment["events"].append({
            "id": normalize_limited_text(item.id, "赛事标识", 40),
            "name": normalize_limited_text(item.name, "赛事名称", 40),
            "kicker": normalize_limited_text(item.kicker, "赛事栏目标签", 60),
            "title": normalize_limited_text(item.title, "赛事介绍标题", 100),
            "description": normalize_limited_text(item.description, "赛事介绍", 1000),
        })
    required_recruitment_values = [value for key, value in recruitment.items() if key != "groups"]
    if not all(required_recruitment_values):
        raise HTTPException(status_code=400, detail="招新栏目文案不能留空")
    if not 1 <= len(payload.recruitment.groups) <= 8:
        raise HTTPException(status_code=400, detail="招新组别需保留 1 至 8 个")
    for item in payload.recruitment.groups:
        name = normalize_limited_text(item.name, "招新组别名称", 40)
        summary = normalize_limited_text(item.summary, "招新组别说明", 240)
        if not name or not summary:
            raise HTTPException(status_code=400, detail="招新组别名称和说明不能留空")
        recruitment["groups"].append({"name": name, "summary": summary})

    timestamp = now_iso()
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO homepage_profile (
                id, team_name, team_intro, stats_json, awards_json, recruitment_json, created_at, updated_at
            )
            VALUES ('profile', ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(id) DO UPDATE SET
                team_name = excluded.team_name,
                team_intro = excluded.team_intro,
                stats_json = excluded.stats_json,
                awards_json = excluded.awards_json,
                recruitment_json = excluded.recruitment_json,
                updated_at = excluded.updated_at
            """,
            (
                team_name,
                team_intro,
                json.dumps(stats, ensure_ascii=False),
                json.dumps(awards, ensure_ascii=False),
                json.dumps(recruitment, ensure_ascii=False),
                timestamp,
                timestamp,
            ),
        )
        row = conn.execute("SELECT * FROM homepage_profile WHERE id = 'profile'").fetchone()
    return homepage_profile_response(row)


def normalize_homepage_order(value: int) -> int:
    if value < 0 or value > 9999:
        raise HTTPException(status_code=400, detail="排序值需在 0 到 9999 之间")
    return value


def normalize_homepage_kind(value: str) -> str:
    kind = value.strip()
    if kind not in HOME_ASSET_KINDS:
        raise HTTPException(status_code=400, detail="媒体类型格式错误")
    return kind


def media_url(filename: str) -> str:
    return f"/api/site-media/{quote(filename)}"


async def save_site_account_photo(account: str, upload: UploadFile) -> dict[str, Any]:
    normalized_account = normalize_account(account)
    if not upload.filename:
        raise HTTPException(status_code=400, detail="请选择个人照片")
    suffix = Path(upload.filename).suffix.lower()
    mime_type = upload.content_type or ""
    if mime_type not in MEMBER_PHOTO_MIME_TYPES or suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
        raise HTTPException(status_code=400, detail="个人照片仅支持 jpg、png、webp")
    content = await upload.read(MAX_MEMBER_PHOTO_LENGTH + 1)
    if not content:
        raise HTTPException(status_code=400, detail="个人照片文件为空")
    if len(content) > MAX_MEMBER_PHOTO_LENGTH:
        raise HTTPException(status_code=413, detail="个人照片不能超过 8MB")

    with db_connection() as conn:
        existing = conn.execute(
            "SELECT account, is_disabled FROM site_account WHERE account = ?",
            (normalized_account,),
        ).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="账号不存在")
        if existing["is_disabled"]:
            raise HTTPException(status_code=403, detail="账号已停用，请联系管理员")

    filename = f"member-{uuid.uuid4().hex}{suffix}"
    (SITE_MEDIA_DIR / filename).write_bytes(content)
    timestamp = now_iso()
    with db_connection() as conn:
        conn.execute(
            "UPDATE site_account SET photo_url = ?, updated_at = ? WHERE account = ?",
            (media_url(filename), timestamp, normalized_account),
        )
        row = conn.execute("SELECT * FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
    return site_account_response(row)


async def save_homepage_award_image(upload: UploadFile) -> dict[str, Any]:
    if not upload.filename:
        raise HTTPException(status_code=400, detail="请选择荣誉图片")
    content = await upload.read()
    if len(content) > MAX_CONTENT_LENGTH:
        raise HTTPException(status_code=413, detail="图片超过 60MB 上限")
    mime_type = upload.content_type or ""
    suffix = Path(upload.filename).suffix.lower()
    if mime_type not in HOME_IMAGE_MIME_TYPES or suffix not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        raise HTTPException(status_code=400, detail="荣誉图片只支持 jpg、png、webp、gif")

    media_id = uuid.uuid4().hex
    filename = f"award-{media_id}{suffix}"
    target_path = SITE_MEDIA_DIR / filename
    target_path.write_bytes(content)
    return {
        "url": media_url(filename),
        "original_filename": upload.filename,
        "mime_type": mime_type,
        "size_bytes": len(content),
    }


async def create_homepage_award(
    upload: UploadFile,
    title_value: str,
    meta_value: str,
    alt_value: str,
    display_order_value: int,
) -> dict[str, Any]:
    title = normalize_limited_text(title_value, "奖项名称", 100)
    meta = normalize_limited_text(meta_value, "奖项说明", 240)
    image_alt = normalize_limited_text(alt_value, "奖项图片说明", 160)
    display_order = normalize_homepage_order(display_order_value)
    if not title:
        raise HTTPException(status_code=400, detail="奖项名称不能为空")
    if not 1 <= display_order <= 12:
        raise HTTPException(status_code=400, detail="荣誉排序需在 1 到 12 之间")

    with db_connection() as conn:
        profile = conn.execute("SELECT * FROM homepage_profile WHERE id = 'profile'").fetchone()
        try:
            awards = ordered_homepage_awards(json.loads(profile["awards_json"] or "[]"))
        except json.JSONDecodeError:
            awards = ordered_homepage_awards(HOME_PROFILE_AWARDS)
        if len(awards) >= 12:
            raise HTTPException(status_code=400, detail="荣誉展示最多支持 12 项")

    uploaded = await save_homepage_award_image(upload)
    awards.insert(min(display_order - 1, len(awards)), {
        "title": title,
        "meta": meta,
        "image_url": uploaded["url"],
        "image_alt": image_alt,
        "display_order": display_order,
    })
    for index, award in enumerate(awards):
        award["display_order"] = index + 1
    timestamp = now_iso()
    with db_connection() as conn:
        conn.execute(
            "UPDATE homepage_profile SET awards_json = ?, updated_at = ? WHERE id = 'profile'",
            (json.dumps(awards, ensure_ascii=False), timestamp),
        )
        profile = conn.execute("SELECT * FROM homepage_profile WHERE id = 'profile'").fetchone()
    return homepage_profile_response(profile)


async def save_homepage_upload(kind: str, upload: UploadFile, alt: str, display_order: int) -> dict[str, Any]:
    normalized_kind = normalize_homepage_kind(kind)
    if not upload.filename:
        raise HTTPException(status_code=400, detail="请选择上传文件")
    mime_type = upload.content_type or ""
    allowed_types = HOME_VIDEO_MIME_TYPES if normalized_kind == "video" else HOME_IMAGE_MIME_TYPES
    if mime_type not in allowed_types:
        raise HTTPException(status_code=400, detail="文件类型不支持")

    suffix = Path(upload.filename).suffix.lower()
    if normalized_kind == "video" and suffix not in {".mp4", ".webm", ".mov"}:
        raise HTTPException(status_code=400, detail="视频只支持 mp4、webm、mov")
    if normalized_kind == "image" and suffix not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        raise HTTPException(status_code=400, detail="图片只支持 jpg、png、webp、gif")

    timestamp = now_iso()
    media_id = uuid.uuid4().hex
    filename = f"{normalized_kind}-{media_id}{suffix}"
    target_path = SITE_MEDIA_DIR / filename
    content_limit = MAX_VIDEO_CONTENT_LENGTH if normalized_kind == "video" else MAX_CONTENT_LENGTH
    size_bytes = 0
    try:
        with target_path.open("wb") as target:
            while chunk := await upload.read(1024 * 1024):
                size_bytes += len(chunk)
                if size_bytes > content_limit:
                    limit_mb = content_limit // 1024 // 1024
                    raise HTTPException(status_code=413, detail=f"文件超过 {limit_mb}MB 上限")
                target.write(chunk)
    except HTTPException:
        target_path.unlink(missing_ok=True)
        raise
    except OSError as exc:
        target_path.unlink(missing_ok=True)
        raise HTTPException(status_code=507, detail="服务器媒体存储空间不足或目录不可写") from exc
    normalized_alt = normalize_limited_text(alt, "媒体说明", 120)
    normalized_order = normalize_homepage_order(display_order)

    try:
        with db_connection() as conn:
            if normalized_kind == "video":
                conn.execute("UPDATE homepage_asset SET is_enabled = 0, updated_at = ? WHERE kind = 'video'", (timestamp,))
            conn.execute(
                """
                INSERT INTO homepage_asset (
                    id, kind, url, storage_path, original_filename, mime_type, size_bytes,
                    alt, display_order, is_enabled, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                """,
                (
                    media_id,
                    normalized_kind,
                    media_url(filename),
                    str(target_path),
                    upload.filename,
                    mime_type,
                    size_bytes,
                    normalized_alt,
                    normalized_order,
                    timestamp,
                    timestamp,
                ),
            )
            row = conn.execute("SELECT * FROM homepage_asset WHERE id = ?", (media_id,)).fetchone()
    except sqlite3.Error as exc:
        target_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail="媒体记录保存失败，请检查数据库状态") from exc
    return homepage_asset_response(row)


def update_homepage_asset(asset_id: str, payload: "HomepageAssetUpdate") -> dict[str, Any]:
    timestamp = now_iso()
    alt = normalize_limited_text(payload.alt, "媒体说明", 120)
    display_order = normalize_homepage_order(payload.display_order)
    with db_connection() as conn:
        existing = conn.execute("SELECT * FROM homepage_asset WHERE id = ?", (asset_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="媒体不存在")
        if existing["kind"] == "video" and payload.is_enabled:
            conn.execute("UPDATE homepage_asset SET is_enabled = 0, updated_at = ? WHERE kind = 'video' AND id <> ?", (timestamp, asset_id))
        conn.execute(
            """
            UPDATE homepage_asset
            SET alt = ?, display_order = ?, is_enabled = ?, updated_at = ?
            WHERE id = ?
            """,
            (alt, display_order, 1 if payload.is_enabled else 0, timestamp, asset_id),
        )
        row = conn.execute("SELECT * FROM homepage_asset WHERE id = ?", (asset_id,)).fetchone()
    return homepage_asset_response(row)


def delete_homepage_asset(asset_id: str) -> dict[str, str]:
    with db_connection() as conn:
        existing = conn.execute("SELECT * FROM homepage_asset WHERE id = ?", (asset_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="媒体不存在")
        conn.execute("DELETE FROM homepage_asset WHERE id = ?", (asset_id,))
    storage_path = existing["storage_path"] or ""
    if storage_path:
        path = Path(storage_path)
        if path.exists() and path.resolve().is_relative_to(SITE_MEDIA_DIR.resolve()):
            path.unlink()
    return {"message": "媒体已删除"}


def create_homepage_quote(payload: "HomepageQuoteCreate") -> dict[str, Any]:
    text = normalize_limited_text(payload.text, "文案", 120)
    if not text:
        raise HTTPException(status_code=400, detail="文案不能为空")
    source = normalize_limited_text(payload.source, "来源", 80)
    display_order = normalize_homepage_order(payload.display_order)
    quote_id = uuid.uuid4().hex
    timestamp = now_iso()
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO homepage_quote (
                id, text, source, display_order, is_enabled, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (quote_id, text, source, display_order, 1 if payload.is_enabled else 0, timestamp, timestamp),
        )
        row = conn.execute("SELECT * FROM homepage_quote WHERE id = ?", (quote_id,)).fetchone()
    return homepage_quote_response(row)


def update_homepage_quote(quote_id: str, payload: "HomepageQuoteUpdate") -> dict[str, Any]:
    text = normalize_limited_text(payload.text, "文案", 120)
    if not text:
        raise HTTPException(status_code=400, detail="文案不能为空")
    source = normalize_limited_text(payload.source, "来源", 80)
    display_order = normalize_homepage_order(payload.display_order)
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute("SELECT id FROM homepage_quote WHERE id = ?", (quote_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="文案不存在")
        conn.execute(
            """
            UPDATE homepage_quote
            SET text = ?, source = ?, display_order = ?, is_enabled = ?, updated_at = ?
            WHERE id = ?
            """,
            (text, source, display_order, 1 if payload.is_enabled else 0, timestamp, quote_id),
        )
        row = conn.execute("SELECT * FROM homepage_quote WHERE id = ?", (quote_id,)).fetchone()
    return homepage_quote_response(row)


def delete_homepage_quote(quote_id: str) -> dict[str, str]:
    with db_connection() as conn:
        existing = conn.execute("SELECT id FROM homepage_quote WHERE id = ?", (quote_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="文案不存在")
        conn.execute("DELETE FROM homepage_quote WHERE id = ?", (quote_id,))
    return {"message": "文案已删除"}


def normalize_danmaku_image_src(value: str) -> str:
    image_src = normalize_limited_text(value, "图片地址", 240)
    if not image_src.startswith(("/", "http://", "https://")):
        raise HTTPException(status_code=400, detail="图片地址格式错误")
    return image_src


def normalize_danmaku_image_key(value: str) -> str:
    return normalize_limited_text(value, "图片标识", 120)


def normalize_danmaku_status(value: str, allow_pending: bool = False) -> str:
    status = value.strip().lower()
    allowed = {"approved", "rejected"}
    if allow_pending:
        allowed.add("pending")
    if status not in allowed:
        raise HTTPException(status_code=400, detail="弹幕审核状态无效")
    return status


def list_homepage_danmaku(image_key: str | None = None) -> dict[str, Any]:
    normalized_image_key = normalize_danmaku_image_key(image_key or "")
    with db_connection() as conn:
        if normalized_image_key:
            rows = conn.execute(
                """
                SELECT
                    danmaku.*,
                    account.full_name AS account_full_name
                FROM homepage_danmaku AS danmaku
                LEFT JOIN site_account AS account
                    ON account.account = danmaku.author_account
                WHERE danmaku.image_key = ? AND danmaku.status = 'approved'
                ORDER BY created_at_ms ASC
                LIMIT 120
                """,
                (normalized_image_key,),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                WITH recent_danmaku AS (
                    SELECT *
                    FROM homepage_danmaku
                    WHERE status = 'approved'
                    ORDER BY created_at_ms DESC
                    LIMIT 240
                )
                SELECT
                    danmaku.*,
                    account.full_name AS account_full_name
                FROM recent_danmaku AS danmaku
                LEFT JOIN site_account AS account
                    ON account.account = danmaku.author_account
                ORDER BY created_at_ms ASC
                """
            ).fetchall()
    return {"messages": [homepage_danmaku_response(row) for row in rows]}


def list_admin_homepage_danmaku() -> dict[str, Any]:
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                danmaku.*,
                account.full_name AS account_full_name
            FROM homepage_danmaku AS danmaku
            LEFT JOIN site_account AS account
                ON account.account = danmaku.author_account
            ORDER BY
                CASE danmaku.status WHEN 'pending' THEN 0 WHEN 'approved' THEN 1 ELSE 2 END,
                danmaku.created_at_ms DESC
            LIMIT 300
            """
        ).fetchall()
    messages = [homepage_danmaku_response(row) for row in rows]
    return {
        "messages": messages,
        "summary": {
            "pending": sum(message["status"] == "pending" for message in messages),
            "approved": sum(message["status"] == "approved" for message in messages),
        },
    }


def review_homepage_danmaku(danmaku_id: str, status_value: str, reviewer: str) -> dict[str, Any]:
    status = normalize_danmaku_status(status_value)
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute("SELECT id FROM homepage_danmaku WHERE id = ?", (danmaku_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="弹幕不存在")
        conn.execute(
            """
            UPDATE homepage_danmaku
            SET status = ?, reviewed_by = ?, reviewed_at = ?
            WHERE id = ?
            """,
            (status, reviewer, timestamp, danmaku_id),
        )
        row = conn.execute("SELECT * FROM homepage_danmaku WHERE id = ?", (danmaku_id,)).fetchone()
    return homepage_danmaku_response(row)


def delete_approved_homepage_danmaku(danmaku_id: str) -> dict[str, str]:
    with db_connection() as conn:
        existing = conn.execute("SELECT status FROM homepage_danmaku WHERE id = ?", (danmaku_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="弹幕不存在")
        if existing["status"] != "approved":
            raise HTTPException(status_code=400, detail="仅已通过弹幕支持删除")
        conn.execute("DELETE FROM homepage_danmaku WHERE id = ?", (danmaku_id,))
    return {"id": danmaku_id}


def create_homepage_danmaku(payload: "HomepageDanmakuCreate") -> dict[str, Any]:
    image_key = normalize_danmaku_image_key(payload.imageKey)
    if not image_key:
        raise HTTPException(status_code=400, detail="图片标识不能为空")
    image_src = normalize_danmaku_image_src(payload.imageSrc)
    text = normalize_limited_text(payload.text, "留言弹幕", 48)
    if not text:
        raise HTTPException(status_code=400, detail="留言弹幕不能为空")
    author_account = normalize_account(payload.authorAccount) if payload.authorAccount.strip() else ""

    timestamp = now_iso()
    created_at_ms = int(time.time() * 1000)
    danmaku_id = uuid.uuid4().hex
    with db_connection() as conn:
        author_name = ""
        if author_account:
            account_row = conn.execute(
                "SELECT account, full_name FROM site_account WHERE account = ?",
                (author_account,),
            ).fetchone()
            if account_row is not None:
                author_account = account_row["account"]
                author_name = account_row["full_name"] or account_row["account"]
            else:
                author_name = author_account
        message_count = conn.execute(
            "SELECT COUNT(*) AS total FROM homepage_danmaku WHERE image_key = ?",
            (image_key,),
        ).fetchone()["total"]
        conn.execute(
            """
            INSERT INTO homepage_danmaku (
                id, image_key, image_src, author_account, author_name, text, track, color, status, created_at_ms, duration, delay, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?, ?, ?)
            """,
            (
                danmaku_id,
                image_key,
                image_src,
                author_account,
                author_name,
                text,
                message_count % DANMAKU_TRACKS,
                DANMAKU_COLORS[message_count % len(DANMAKU_COLORS)],
                created_at_ms,
                8 + (message_count % 5),
                0,
                timestamp,
            ),
        )
        row = conn.execute("SELECT * FROM homepage_danmaku WHERE id = ?", (danmaku_id,)).fetchone()
    return {"message": homepage_danmaku_response(row)}


class LoginRequest(BaseModel):
    password: str
    role: str = "admin"


class SiteAccountProfile(BaseModel):
    full_name: str = ""
    gender: str = ""
    grade: str = ""
    member_status: str = ""
    permission_level: str = ""
    department: str = ""
    cohort: str = ""
    role: str = ""
    phone: str = ""
    email: str = ""
    bio: str = ""
    photo_url: str = ""


class SiteAccountRequest(BaseModel):
    account: str
    password: str


class SiteAccountRegisterRequest(SiteAccountRequest, SiteAccountProfile):
    pass


class Image2AccessRequest(BaseModel):
    image2_allowed: bool


class RewardScoreRequest(BaseModel):
    reward_score: int


class SiteAccountAdminUpdate(SiteAccountProfile):
    reward_score: int = 0
    image2_allowed: bool = False
    is_disabled: bool = False
    admin_note: str = ""


class SiteAccountPasswordResetRequest(BaseModel):
    new_password: str


class ForumPostCreateRequest(BaseModel):
    title: str
    content: str
    author_account: str


class ForumPostUpdateRequest(BaseModel):
    title: str
    content: str
    author_account: str


class ForumReplyCreateRequest(BaseModel):
    content: str
    author_account: str


class ForumReplyUpdateRequest(BaseModel):
    content: str
    author_account: str


class ForumModerationRequest(BaseModel):
    status: str
    reject_reason: str = ""
    is_pinned: bool | None = None
    is_locked: bool | None = None


class SSLInterviewApplicationRequest(BaseModel):
    applicant_account: str
    self_intro: str
    interview_direction: str
    interview_time: str


class SSLInterviewReviewRequest(BaseModel):
    status: str
    interview_location: str = ""
    rejection_reason: str = ""


class FleaMarketItemCreateRequest(BaseModel):
    name: str
    image_src: str = ""
    location: str
    summary: str
    detail: str
    contact: str
    tags: str = ""
    author_account: str


class FleaMarketItemUpdateRequest(FleaMarketItemCreateRequest):
    pass


class FleaMarketStatusRequest(BaseModel):
    author_account: str
    status: str


class SeasonPlanItem(BaseModel):
    robot_type: str
    task_title: str
    status: str
    target: str
    assignee_account: str = ""
    is_completed: bool = False


class SeasonPlanRequest(BaseModel):
    season_year: int
    month: int
    editor_account: str
    plans: list[SeasonPlanItem]


class HomepageAssetUpdate(BaseModel):
    alt: str = ""
    display_order: int = 0
    is_enabled: bool = True


class HomepageRecruitmentBannerUpdate(BaseModel):
    text: str
    action_text: str
    is_enabled: bool = True


class HomepageStatItem(BaseModel):
    value: str
    label: str


class HomepageAwardItem(BaseModel):
    title: str
    meta: str = ""
    image_url: str = ""
    image_alt: str = ""
    display_order: int = 0


class HomepageRecruitmentGroupItem(BaseModel):
    name: str
    summary: str


class HomepageRecruitmentEventItem(BaseModel):
    id: str
    name: str
    kicker: str
    title: str
    description: str


class HomepageRecruitmentContent(BaseModel):
    season_label: str
    title: str
    intro: str
    event_kicker: str
    event_title: str
    event_description: str
    events: list[HomepageRecruitmentEventItem] = []
    groups_kicker: str
    groups_title: str
    groups: list[HomepageRecruitmentGroupItem]
    qr_text: str


class HomepageProfileUpdate(BaseModel):
    team_name: str
    team_intro: str
    stats: list[HomepageStatItem]
    awards: list[HomepageAwardItem]
    recruitment: HomepageRecruitmentContent


class HomepageQuoteCreate(BaseModel):
    text: str
    source: str = ""
    display_order: int = 0
    is_enabled: bool = True


class HomepageQuoteUpdate(HomepageQuoteCreate):
    pass


class HomepageDanmakuCreate(BaseModel):
    imageKey: str = ""
    imageSrc: str
    text: str
    authorAccount: str = ""


class HomepageDanmakuReview(BaseModel):
    status: str


class RecruitmentQuestionCreate(BaseModel):
    author_account: str = ""
    content: str = ""


class RecruitmentFaqSave(BaseModel):
    question: str
    answer: str
    display_order: int = 0
    is_enabled: bool = True


class RowIdsRequest(BaseModel):
    row_ids: list[str]
    note: str = ""


class ReimburseRequest(BaseModel):
    record_ids: list[str]


class AgentThread:
    def __init__(self) -> None:
        self._stop_event = threading.Event()
        self._thread = threading.Thread(target=self._loop, daemon=True, name="local-review")

    def start(self) -> None:
        if not self._thread.is_alive():
            self._thread.start()

    def _loop(self) -> None:
        while not self._stop_event.is_set():
            try:
                run_agent_once()
            except Exception as exc:  # noqa: BLE001
                log_process("local_review", "error", f"后台任务异常：{exc}")
            self._stop_event.wait(AGENT_INTERVAL_SECONDS)


init_db()
agent_thread = AgentThread()
agent_thread.start()

app = FastAPI(title="PRINTK 团队门户 API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[origin.strip() for origin in FRONTEND_ORIGIN.split(",") if origin.strip()],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def write_api_gate(request: Request, call_next):
    if request.method in {"POST", "PUT", "PATCH", "DELETE"} and (SITE_MODE == "test" or not ENABLE_WRITE_API):
        return JSONResponse({"detail": "feature disabled"}, status_code=503)
    return await call_next(request)


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok", "time": now_iso()}


@app.get("/api/homepage")
def homepage_content() -> dict[str, Any]:
    return get_homepage_content(include_disabled=False)


@app.post("/api/homepage/recruitment-questions", status_code=201)
def create_recruitment_question_route(payload: RecruitmentQuestionCreate) -> dict[str, Any]:
    return create_recruitment_question(payload)


@app.get("/api/homepage/danmaku")
def homepage_danmaku(image_key: str | None = Query(default=None)) -> dict[str, Any]:
    return list_homepage_danmaku(image_key)


@app.post("/api/homepage/danmaku")
def create_homepage_danmaku_route(payload: HomepageDanmakuCreate) -> dict[str, Any]:
    return create_homepage_danmaku(payload)


@app.get("/api/admin/homepage/danmaku")
def admin_homepage_danmaku(_: str = Depends(require_admin)) -> dict[str, Any]:
    return list_admin_homepage_danmaku()


@app.put("/api/admin/homepage/danmaku/{danmaku_id}")
def review_homepage_danmaku_route(
    danmaku_id: str,
    payload: HomepageDanmakuReview,
    reviewer: str = Depends(require_admin),
) -> dict[str, Any]:
    return review_homepage_danmaku(danmaku_id, payload.status, reviewer)


@app.delete("/api/admin/homepage/danmaku/{danmaku_id}")
def delete_homepage_danmaku_route(
    danmaku_id: str,
    _: str = Depends(require_admin),
) -> dict[str, str]:
    return delete_approved_homepage_danmaku(danmaku_id)


@app.get("/api/site-media/{filename}")
def get_site_media(filename: str) -> FileResponse:
    path = (SITE_MEDIA_DIR / filename).resolve()
    if not path.is_file() or not path.is_relative_to(SITE_MEDIA_DIR.resolve()):
        raise HTTPException(status_code=404, detail="媒体文件不存在")
    return FileResponse(path)


@app.get("/api/admin/homepage")
def admin_homepage_content(_: str = Depends(require_admin)) -> dict[str, Any]:
    return get_homepage_content(include_disabled=True)


@app.post("/api/admin/homepage/faqs", status_code=201)
def create_recruitment_faq_route(payload: RecruitmentFaqSave, _: str = Depends(require_admin)) -> dict[str, Any]:
    return save_recruitment_faq(payload)


@app.put("/api/admin/homepage/faqs/{faq_id}")
def update_recruitment_faq_route(faq_id: str, payload: RecruitmentFaqSave, _: str = Depends(require_admin)) -> dict[str, Any]:
    return save_recruitment_faq(payload, faq_id)


@app.delete("/api/admin/homepage/faqs/{faq_id}")
def delete_recruitment_faq_route(faq_id: str, _: str = Depends(require_admin)) -> dict[str, str]:
    return delete_recruitment_faq(faq_id)


@app.delete("/api/admin/homepage/recruitment-questions/{question_id}")
def delete_recruitment_question_route(question_id: str, _: str = Depends(require_admin)) -> dict[str, str]:
    return delete_recruitment_question(question_id)


@app.put("/api/admin/homepage/recruitment-banner")
def update_homepage_recruitment_banner_route(
    payload: HomepageRecruitmentBannerUpdate,
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    return update_homepage_recruitment_banner(payload)


@app.put("/api/admin/homepage/campus-banner")
def update_homepage_campus_banner_route(
    payload: HomepageRecruitmentBannerUpdate,
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    return update_homepage_campus_banner(payload)


@app.put("/api/admin/homepage/profile")
def update_homepage_profile_route(
    payload: HomepageProfileUpdate,
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    return update_homepage_profile(payload)


@app.post("/api/admin/homepage/assets")
async def upload_homepage_asset(
    kind: str = Form(...),
    alt: str = Form(""),
    display_order: int = Form(0),
    file: UploadFile = File(...),
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    return await save_homepage_upload(kind, file, alt, display_order)


@app.post("/api/admin/homepage/award-image")
async def upload_homepage_award_image(
    file: UploadFile = File(...),
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    return await save_homepage_award_image(file)


@app.post("/api/admin/homepage/awards")
async def create_homepage_award_route(
    title: str = Form(...),
    meta: str = Form(""),
    image_alt: str = Form(""),
    display_order: int = Form(...),
    file: UploadFile = File(...),
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    return await create_homepage_award(file, title, meta, image_alt, display_order)


@app.put("/api/admin/homepage/assets/{asset_id}")
def update_homepage_asset_route(
    asset_id: str,
    payload: HomepageAssetUpdate,
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    return update_homepage_asset(asset_id, payload)


@app.delete("/api/admin/homepage/assets/{asset_id}")
def delete_homepage_asset_route(asset_id: str, _: str = Depends(require_admin)) -> dict[str, str]:
    return delete_homepage_asset(asset_id)


@app.post("/api/admin/homepage/quotes")
def create_homepage_quote_route(
    payload: HomepageQuoteCreate,
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    return create_homepage_quote(payload)


@app.put("/api/admin/homepage/quotes/{quote_id}")
def update_homepage_quote_route(
    quote_id: str,
    payload: HomepageQuoteUpdate,
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    return update_homepage_quote(quote_id, payload)


@app.delete("/api/admin/homepage/quotes/{quote_id}")
def delete_homepage_quote_route(quote_id: str, _: str = Depends(require_admin)) -> dict[str, str]:
    return delete_homepage_quote(quote_id)


@app.post("/api/site-accounts/register")
def register_site_account(payload: SiteAccountRegisterRequest) -> dict[str, str]:
    account = normalize_account(payload.account)
    password = validate_site_password(payload.password)
    profile = normalize_site_profile(payload)
    profile["member_status"] = "非战队队员"
    profile["permission_level"] = "普通队员"
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute("SELECT account FROM site_account WHERE account = ?", (account,)).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail="账号已存在")
        conn.execute(
            """
            INSERT INTO site_account (
                account, password_hash, full_name, gender, grade, member_status, permission_level,
                department, cohort, role, phone, email, bio, photo_url, reward_score, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
            """,
            (
                account,
                hash_site_password(password),
                profile["full_name"],
                profile["gender"],
                profile["grade"],
                profile["member_status"],
                profile["permission_level"] or "普通队员",
                profile["department"],
                profile["cohort"],
                profile["role"],
                profile["phone"],
                profile["email"],
                profile["bio"],
                profile["photo_url"],
                timestamp,
                timestamp,
            ),
        )
    return {"account": account}


@app.post("/api/site-accounts/login")
def login_site_account(payload: SiteAccountRequest) -> dict[str, str]:
    account = normalize_account(payload.account)
    with db_connection() as conn:
        row = conn.execute("SELECT * FROM site_account WHERE account = ?", (account,)).fetchone()
        if row and not row["is_disabled"] and verify_site_password(payload.password, row["password_hash"]):
            conn.execute(
                "UPDATE site_account SET last_login_at = ?, updated_at = ? WHERE account = ?",
                (now_iso(), now_iso(), account),
            )
            return {"account": account}
    if not row or not verify_site_password(payload.password, row["password_hash"]):
        raise HTTPException(status_code=401, detail="账号或密码错误")
    raise HTTPException(status_code=403, detail="账号已停用，请联系管理员")


@app.get("/api/site-accounts/{account}")
def get_site_account(account: str) -> dict[str, Any]:
    normalized_account = normalize_account(account)
    with db_connection() as conn:
        row = conn.execute("SELECT * FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="账号不存在")
    return site_account_response(row)


@app.put("/api/site-accounts/{account}")
def update_site_account(account: str, payload: SiteAccountProfile) -> dict[str, Any]:
    normalized_account = normalize_account(account)
    profile = normalize_site_profile(payload)
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute(
            "SELECT account, is_disabled, member_status, permission_level FROM site_account WHERE account = ?",
            (normalized_account,),
        ).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="账号不存在")
        if existing["is_disabled"]:
            raise HTTPException(status_code=403, detail="账号已停用，请联系管理员")
        conn.execute(
            """
            UPDATE site_account
            SET full_name = ?, gender = ?, grade = ?, member_status = ?, permission_level = ?,
                department = ?, cohort = ?, role = ?, phone = ?, email = ?, bio = ?, photo_url = ?,
                reward_score = CASE WHEN ? IN ('正式队员', '老队员') THEN reward_score ELSE 0 END,
                updated_at = ?
            WHERE account = ?
            """,
            (
                profile["full_name"],
                profile["gender"],
                profile["grade"],
                existing["member_status"] or "非战队队员",
                existing["permission_level"] or "普通队员",
                profile["department"],
                profile["cohort"],
                profile["role"],
                profile["phone"],
                profile["email"],
                profile["bio"],
                profile["photo_url"],
                existing["member_status"] or "非战队队员",
                timestamp,
                normalized_account,
            ),
        )
        row = conn.execute("SELECT * FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
    return site_account_response(row)


@app.post("/api/site-accounts/{account}/photo")
async def upload_site_account_photo(account: str, file: UploadFile = File(...)) -> dict[str, Any]:
    return await save_site_account_photo(account, file)


@app.post("/api/ssl/applications", status_code=201)
def submit_ssl_interview_application(payload: SSLInterviewApplicationRequest) -> dict[str, Any]:
    self_intro = normalize_limited_text(payload.self_intro, "自我简介", 1000)
    if not self_intro:
        raise HTTPException(status_code=400, detail="请填写自我简介")
    direction = normalize_choice(payload.interview_direction, "面试方向", SSL_INTERVIEW_DIRECTIONS)
    interview_time = normalize_limited_text(payload.interview_time, "面试时间", 32)
    try:
        parsed_time = datetime.fromisoformat(interview_time)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="面试时间格式错误") from exc
    if parsed_time <= datetime.now():
        raise HTTPException(status_code=400, detail="请选择当前时间之后的面试时间")

    timestamp = now_iso()
    with db_connection() as conn:
        account_row = require_active_site_account(conn, payload.applicant_account)
        if not (account_row["full_name"] or "").strip() or not (account_row["grade"] or "").strip():
            raise HTTPException(status_code=400, detail="请先在个人中心完善姓名和年级")
        existing = conn.execute(
            "SELECT id, status FROM ssl_interview_application WHERE applicant_account = ?",
            (account_row["account"],),
        ).fetchone()
        if existing and existing["status"] in {"pending", "approved"}:
            raise HTTPException(status_code=409, detail="当前已有待审核或已通过的面试申请")
        if existing:
            application_id = existing["id"]
            conn.execute(
                """
                UPDATE ssl_interview_application
                SET self_intro = ?, interview_direction = ?, interview_time = ?, status = 'pending',
                    interview_location = '', rejection_reason = '', reviewed_by = '', reviewed_at = '', updated_at = ?
                WHERE id = ?
                """,
                (self_intro, direction, parsed_time.isoformat(timespec="minutes"), timestamp, application_id),
            )
            conn.execute(
                "DELETE FROM site_message WHERE category = 'ssl_interview' AND related_id = ?",
                (application_id,),
            )
        else:
            application_id = uuid.uuid4().hex
            conn.execute(
                """
                INSERT INTO ssl_interview_application (
                    id, applicant_account, self_intro, interview_direction, interview_time,
                    status, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, 'pending', ?, ?)
                """,
                (
                    application_id,
                    account_row["account"],
                    self_intro,
                    direction,
                    parsed_time.isoformat(timespec="minutes"),
                    timestamp,
                    timestamp,
                ),
            )
        row = conn.execute(
            """
            SELECT application.*, account.full_name, account.grade
            FROM ssl_interview_application AS application
            JOIN site_account AS account ON account.account = application.applicant_account
            WHERE application.id = ?
            """,
            (application_id,),
        ).fetchone()
    return ssl_application_response(row)


@app.get("/api/ssl/applications/{account}")
def get_ssl_interview_application(account: str) -> dict[str, Any]:
    with db_connection() as conn:
        account_row = require_active_site_account(conn, account)
        row = conn.execute(
            """
            SELECT application.*, account.full_name, account.grade
            FROM ssl_interview_application AS application
            JOIN site_account AS account ON account.account = application.applicant_account
            WHERE application.applicant_account = ?
            """,
            (account_row["account"],),
        ).fetchone()
    return {"application": ssl_application_response(row) if row else None}


@app.get("/api/site-messages/{account}")
def list_site_messages(account: str, limit: int = Query(default=50, ge=1, le=100)) -> dict[str, Any]:
    with db_connection() as conn:
        account_row = require_active_site_account(conn, account)
        rows = conn.execute(
            """
            SELECT id, category, title, content, related_id, read_at, created_at
            FROM site_message
            WHERE recipient_account = ?
            ORDER BY created_at DESC
            LIMIT ?
            """,
            (account_row["account"], limit),
        ).fetchall()
        unread_count = conn.execute(
            "SELECT COUNT(*) FROM site_message WHERE recipient_account = ? AND COALESCE(read_at, '') = ''",
            (account_row["account"],),
        ).fetchone()[0]
    return {"messages": [site_message_response(row) for row in rows], "unread_count": unread_count}


@app.put("/api/site-messages/{account}/read")
def mark_site_messages_read(account: str) -> dict[str, int]:
    with db_connection() as conn:
        account_row = require_active_site_account(conn, account)
        cursor = conn.execute(
            """
            UPDATE site_message
            SET read_at = ?
            WHERE recipient_account = ? AND COALESCE(read_at, '') = ''
            """,
            (now_iso(), account_row["account"]),
        )
    return {"marked_read": cursor.rowcount, "unread_count": 0}


@app.get("/api/admin/ssl/applications")
def list_ssl_interview_applications(
    status: str = Query(default=""),
    limit: int = Query(default=20, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    normalized_status = normalize_choice(status, "申请状态", SSL_APPLICATION_STATUSES | {""})
    where_sql = "WHERE application.status = ?" if normalized_status else ""
    parameters: list[Any] = [normalized_status] if normalized_status else []
    with db_connection() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) FROM ssl_interview_application AS application {where_sql}",
            parameters,
        ).fetchone()[0]
        rows = conn.execute(
            f"""
            SELECT application.*, account.full_name, account.grade
            FROM ssl_interview_application AS application
            JOIN site_account AS account ON account.account = application.applicant_account
            {where_sql}
            ORDER BY CASE application.status WHEN 'pending' THEN 0 ELSE 1 END,
                     application.interview_time, application.created_at DESC
            LIMIT ? OFFSET ?
            """,
            [*parameters, limit, offset],
        ).fetchall()
    return {"applications": [ssl_application_response(row) for row in rows], "total": total}


@app.put("/api/admin/ssl/applications/{application_id}")
def review_ssl_interview_application(
    application_id: str,
    payload: SSLInterviewReviewRequest,
    reviewer: str = Depends(require_admin),
) -> dict[str, Any]:
    status = normalize_choice(payload.status, "审核状态", {"approved", "rejected"})
    location = normalize_limited_text(payload.interview_location, "面试地点", 120)
    reason = normalize_limited_text(payload.rejection_reason, "未通过原因", 500)
    if status == "approved" and not location:
        raise HTTPException(status_code=400, detail="通过申请时请填写面试地点")
    if status == "rejected" and not reason:
        raise HTTPException(status_code=400, detail="拒绝申请时请填写未通过原因")

    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute(
            "SELECT * FROM ssl_interview_application WHERE id = ?",
            (application_id,),
        ).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="面试申请不存在")
        conn.execute(
            """
            UPDATE ssl_interview_application
            SET status = ?, interview_location = ?, rejection_reason = ?,
                reviewed_by = ?, reviewed_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, location if status == "approved" else "", reason if status == "rejected" else "", reviewer, timestamp, timestamp, application_id),
        )
        if status == "approved":
            title = "SSL 部面试申请已通过"
            content = f"面试时间：{existing['interview_time'].replace('T', ' ')}；面试地点：{location}。请按时参加。"
        else:
            title = "SSL 部面试申请结果"
            content = f"本次申请未通过。原因：{reason}"
        conn.execute("DELETE FROM site_message WHERE category = 'ssl_interview' AND related_id = ?", (application_id,))
        conn.execute(
            """
            INSERT INTO site_message (id, recipient_account, category, title, content, related_id, created_at)
            VALUES (?, ?, 'ssl_interview', ?, ?, ?, ?)
            """,
            (uuid.uuid4().hex, existing["applicant_account"], title, content, application_id, timestamp),
        )
        row = conn.execute(
            """
            SELECT application.*, account.full_name, account.grade
            FROM ssl_interview_application AS application
            JOIN site_account AS account ON account.account = application.applicant_account
            WHERE application.id = ?
            """,
            (application_id,),
        ).fetchone()
    return ssl_application_response(row)


@app.get("/api/admin/ssl/applications/export.csv")
def export_ssl_interview_applications(_: str = Depends(require_admin)) -> StreamingResponse:
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT account.full_name, account.grade, application.interview_direction, application.interview_time
            FROM ssl_interview_application AS application
            JOIN site_account AS account ON account.account = application.applicant_account
            ORDER BY application.interview_time, application.created_at
            """
        ).fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["姓名", "年级", "面试方向", "面试时间"])
    writer.writerows(
        [row["full_name"], row["grade"], row["interview_direction"], row["interview_time"].replace("T", " ")]
        for row in rows
    )
    content = ("\ufeff" + output.getvalue()).encode("utf-8")
    filename = quote("SSL部面试申请.csv")
    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


PUBLIC_MEMBER_STATUSES = {"梯队队员", "正式队员", "老队员", "退役队员"}
RETIRED_MEMBER_STATUSES = {"老队员", "退役队员"}


def member_response(row: sqlite3.Row) -> dict[str, Any]:
    member_status = row["member_status"] or ""
    membership_state = "retired" if member_status in RETIRED_MEMBER_STATUSES else "active"
    return {
        "id": row["account"],
        "account": row["account"],
        "name": row["full_name"] or row["account"],
        "membership_state": membership_state,
        "member_status": member_status,
        "cohort": row["cohort"] or "",
        "group": row["department"] or "",
        "role": row["role"] or row["permission_level"] or "",
        "grade": row["grade"] or "",
        "bio": row["bio"] or "",
        "photo_url": row["photo_url"] or "",
        "updated_at": row["updated_at"],
    }


@app.get("/api/members")
def list_public_members() -> dict[str, Any]:
    placeholders = ",".join("?" for _ in PUBLIC_MEMBER_STATUSES)
    with db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT account, full_name, grade, member_status, permission_level,
                department, cohort, role, bio, photo_url, updated_at
            FROM site_account
            WHERE is_disabled = 0
                AND member_status IN ({placeholders})
                AND COALESCE(full_name, '') <> ''
            ORDER BY
                CASE WHEN member_status IN ('老队员', '退役队员') THEN 1 ELSE 0 END,
                cohort DESC,
                department ASC,
                updated_at DESC,
                account ASC
            LIMIT 300
            """,
            tuple(sorted(PUBLIC_MEMBER_STATUSES)),
        ).fetchall()
    return {"members": [member_response(row) for row in rows]}


@app.get("/api/admin/site-accounts")
def list_site_accounts(
    keyword: str = Query(default="", max_length=80),
    member_status: str = Query(default=""),
    permission_level: str = Query(default=""),
    department: str = Query(default=""),
    state: str = Query(default="all"),
    image2: str = Query(default="all"),
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    normalized_member_status = normalize_choice(member_status, "身份信息", MEMBER_STATUS_OPTIONS)
    normalized_permission_level = normalize_choice(permission_level, "权限", PERMISSION_LEVEL_OPTIONS)
    normalized_department = normalize_choice(department, "部门信息", DEPARTMENT_OPTIONS)
    normalized_keyword = keyword.strip()
    if state not in {"all", "enabled", "disabled"}:
        raise HTTPException(status_code=400, detail="账号状态格式错误")
    if image2 not in {"all", "allowed", "denied"}:
        raise HTTPException(status_code=400, detail="图片工具权限格式错误")

    where = ["1 = 1"]
    params: list[Any] = []
    if normalized_keyword:
        where.append("(account LIKE ? OR full_name LIKE ? OR phone LIKE ? OR email LIKE ? OR cohort LIKE ? OR role LIKE ?)")
        like_value = f"%{normalized_keyword}%"
        params.extend([like_value, like_value, like_value, like_value, like_value, like_value])
    if normalized_member_status:
        where.append("member_status = ?")
        params.append(normalized_member_status)
    if normalized_permission_level:
        where.append("permission_level = ?")
        params.append(normalized_permission_level)
    if normalized_department:
        where.append("department = ?")
        params.append(normalized_department)
    if state == "enabled":
        where.append("is_disabled = 0")
    if state == "disabled":
        where.append("is_disabled = 1")
    if image2 == "allowed":
        where.append("image2_allowed = 1")
    if image2 == "denied":
        where.append("image2_allowed = 0")

    with db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT *
            FROM site_account
            WHERE {' AND '.join(where)}
            ORDER BY is_disabled ASC, updated_at DESC, account ASC
            LIMIT 200
            """,
            params,
        ).fetchall()
        summary = conn.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN is_disabled = 0 THEN 1 ELSE 0 END) AS enabled,
                SUM(CASE WHEN is_disabled = 1 THEN 1 ELSE 0 END) AS disabled,
                SUM(CASE WHEN image2_allowed = 1 THEN 1 ELSE 0 END) AS image2_allowed
            FROM site_account
            """
        ).fetchone()
    return {
        "accounts": [site_account_response(row, include_admin=True) for row in rows],
        "summary": {
            "total": summary["total"] or 0,
            "enabled": summary["enabled"] or 0,
            "disabled": summary["disabled"] or 0,
            "image2_allowed": summary["image2_allowed"] or 0,
        },
    }


@app.put("/api/admin/site-accounts/{account}")
def update_site_account_by_admin(
    account: str,
    payload: SiteAccountAdminUpdate,
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    normalized_account = normalize_account(account)
    profile = normalize_site_profile(payload)
    admin_note = normalize_limited_text(payload.admin_note, "后台备注", 200)
    reward_score = normalize_reward_score(payload.reward_score) if reward_eligible(profile["member_status"]) else 0
    if payload.image2_allowed and profile["member_status"] != "正式队员":
        raise HTTPException(status_code=400, detail="图片工具权限只支持正式队员账号")
    if payload.is_disabled and payload.image2_allowed:
        raise HTTPException(status_code=400, detail="停用账号不能添加图片工具权限")
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute("SELECT account FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="账号不存在")
        conn.execute(
            """
            UPDATE site_account
            SET full_name = ?, gender = ?, grade = ?, member_status = ?, permission_level = ?,
                department = ?, cohort = ?, role = ?, phone = ?, email = ?, bio = ?, photo_url = ?,
                reward_score = ?, image2_allowed = ?, is_disabled = ?, admin_note = ?, updated_at = ?
            WHERE account = ?
            """,
            (
                profile["full_name"],
                profile["gender"],
                profile["grade"],
                profile["member_status"],
                profile["permission_level"] or "普通队员",
                profile["department"],
                profile["cohort"],
                profile["role"],
                profile["phone"],
                profile["email"],
                profile["bio"],
                profile["photo_url"],
                reward_score,
                1 if payload.image2_allowed else 0,
                1 if payload.is_disabled else 0,
                admin_note,
                timestamp,
                normalized_account,
            ),
        )
        account_admin_log(conn, normalized_account, "update", "管理员更新账号资料")
        updated = conn.execute("SELECT * FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
    return site_account_response(updated, include_admin=True)


@app.delete("/api/admin/site-accounts/{account}")
def delete_site_account_by_admin(
    account: str,
    _: str = Depends(require_admin),
) -> dict[str, str]:
    normalized_account = normalize_account(account)
    with db_connection() as conn:
        existing = conn.execute("SELECT account FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="账号不存在")
        forum_post_count = conn.execute(
            "SELECT COUNT(*) AS total FROM forum_post WHERE author_account = ?",
            (normalized_account,),
        ).fetchone()["total"]
        forum_reply_count = conn.execute(
            "SELECT COUNT(*) AS total FROM forum_reply WHERE author_account = ?",
            (normalized_account,),
        ).fetchone()["total"]
        market_item_count = conn.execute(
            "SELECT COUNT(*) AS total FROM flea_market_item WHERE author_account = ?",
            (normalized_account,),
        ).fetchone()["total"]
        if forum_post_count or forum_reply_count or market_item_count:
            raise HTTPException(status_code=400, detail="账号存在前台内容，请先停用账号")
        conn.execute("DELETE FROM site_account_admin_log WHERE account = ?", (normalized_account,))
        conn.execute("UPDATE season_plan SET assignee_account = '' WHERE assignee_account = ?", (normalized_account,))
        conn.execute("UPDATE homepage_danmaku SET author_account = '' WHERE author_account = ?", (normalized_account,))
        conn.execute("DELETE FROM site_account WHERE account = ?", (normalized_account,))
    return {"account": normalized_account, "message": "账号已删除"}


@app.put("/api/admin/site-accounts/{account}/reward-score")
def update_site_account_reward_score(
    account: str,
    payload: RewardScoreRequest,
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    normalized_account = normalize_account(account)
    reward_score = normalize_reward_score(payload.reward_score)
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute(
            "SELECT account, member_status FROM site_account WHERE account = ?",
            (normalized_account,),
        ).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="账号不存在")
        if not reward_eligible(existing["member_status"] or ""):
            raise HTTPException(status_code=400, detail="该账号身份不支持奖励分")
        conn.execute(
            """
            UPDATE site_account
            SET reward_score = ?, updated_at = ?
            WHERE account = ?
            """,
            (reward_score, timestamp, normalized_account),
        )
        account_admin_log(conn, normalized_account, "reward_score", "管理员更新奖励分")
        updated = conn.execute("SELECT * FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
    return site_account_response(updated, include_admin=True)


@app.get("/api/reward-ranking")
def reward_ranking() -> dict[str, Any]:
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT account, full_name, member_status, department, reward_score, updated_at
            FROM site_account
            WHERE is_disabled = 0
                AND member_status IN ('正式队员', '老队员')
            ORDER BY reward_score DESC, updated_at ASC, account ASC
            LIMIT 200
            """
        ).fetchall()
    return {
        "ranking": [
            {
                "rank": index,
                "account": row["account"],
                "full_name": row["full_name"] or row["account"],
                "member_status": row["member_status"] or "",
                "department": row["department"] or "",
                "reward_score": row["reward_score"] or 0,
                "updated_at": row["updated_at"] or "",
            }
            for index, row in enumerate(rows, start=1)
        ]
    }


@app.post("/api/admin/site-accounts/{account}/reset-password")
def reset_site_account_password(
    account: str,
    payload: SiteAccountPasswordResetRequest,
    _: str = Depends(require_admin),
) -> dict[str, str]:
    normalized_account = normalize_account(account)
    new_password = validate_site_password(payload.new_password)
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute("SELECT account FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="账号不存在")
        conn.execute(
            "UPDATE site_account SET password_hash = ?, updated_at = ? WHERE account = ?",
            (hash_site_password(new_password), timestamp, normalized_account),
        )
        account_admin_log(conn, normalized_account, "reset_password", "管理员重置账号密码")
    return {"account": normalized_account, "message": "密码已重置"}


@app.get("/api/admin/site-accounts/image2-access")
def list_image2_access_accounts(_: str = Depends(require_admin)) -> dict[str, Any]:
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM site_account
            WHERE member_status = ? AND is_disabled = 0
            ORDER BY image2_allowed DESC, updated_at DESC, account ASC
            """,
            ("正式队员",),
        ).fetchall()
    return {"accounts": [site_account_response(row) for row in rows]}


@app.put("/api/admin/site-accounts/{account}/image2-access")
def update_image2_access(
    account: str,
    payload: Image2AccessRequest,
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    normalized_account = normalize_account(account)
    timestamp = now_iso()
    with db_connection() as conn:
        row = conn.execute("SELECT * FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="账号不存在")
        if row["is_disabled"]:
            raise HTTPException(status_code=400, detail="停用账号不能添加图片工具权限")
        if payload.image2_allowed and row["member_status"] != "正式队员":
            raise HTTPException(status_code=400, detail="图片工具权限只支持正式队员账号")
        conn.execute(
            """
            UPDATE site_account
            SET image2_allowed = ?, updated_at = ?
            WHERE account = ?
            """,
            (1 if payload.image2_allowed else 0, timestamp, normalized_account),
        )
        account_admin_log(conn, normalized_account, "image2_access", "管理员更新图片工具权限")
        updated = conn.execute("SELECT * FROM site_account WHERE account = ?", (normalized_account,)).fetchone()
    return site_account_response(updated, include_admin=True)


@app.get("/api/admin/forum")
def list_forum_management(
    status: str = Query(default="all"),
    _: str = Depends(require_admin),
) -> dict[str, Any]:
    status_filter = status.strip()
    if status_filter != "all":
        status_filter = normalize_forum_status(status_filter)
    where_sql = ""
    params: list[Any] = [FORUM_PUBLIC_STATUS]
    if status_filter != "all":
        where_sql = "WHERE post.status = ?"
        params.append(status_filter)
    with db_connection() as conn:
        posts = conn.execute(
            f"""
            SELECT
                post.*,
                account.full_name,
                COUNT(reply.id) AS reply_count
            FROM forum_post AS post
            LEFT JOIN site_account AS account ON account.account = post.author_account
            LEFT JOIN forum_reply AS reply
                ON reply.post_id = post.id
                AND reply.status = ?
                AND COALESCE(reply.deleted_at, '') = ''
            {where_sql}
            GROUP BY post.id
            ORDER BY
                CASE WHEN post.status = 'pending' THEN 0 ELSE 1 END,
                post.is_pinned DESC,
                post.created_at DESC
            LIMIT 200
            """,
            params,
        ).fetchall()
        replies = conn.execute(
            """
            SELECT
                reply.*,
                account.full_name,
                post.title AS post_title
            FROM forum_reply AS reply
            LEFT JOIN site_account AS account ON account.account = reply.author_account
            LEFT JOIN forum_post AS post ON post.id = reply.post_id
            ORDER BY
                CASE WHEN reply.status = 'pending' THEN 0 ELSE 1 END,
                reply.created_at DESC
            LIMIT 200
            """
        ).fetchall()
    reply_payload = []
    for row in replies:
        item = forum_reply_response(row)
        item["post_title"] = row["post_title"] or ""
        reply_payload.append(item)
    return {
        "posts": [forum_post_response(row) for row in posts],
        "replies": reply_payload,
    }


@app.put("/api/admin/forum/posts/{post_id}")
def moderate_forum_post(
    post_id: str,
    payload: ForumModerationRequest,
    reviewer: str = Depends(require_admin),
) -> dict[str, Any]:
    status = normalize_forum_status(payload.status)
    reject_reason = normalize_forum_content(payload.reject_reason, "处理说明", 500) if payload.reject_reason.strip() else ""
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute("SELECT id FROM forum_post WHERE id = ?", (post_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="帖子不存在")
        conn.execute(
            """
            UPDATE forum_post
            SET status = ?,
                reject_reason = ?,
                reviewed_by = ?,
                reviewed_at = ?,
                is_pinned = COALESCE(?, is_pinned),
                is_locked = COALESCE(?, is_locked),
                updated_at = ?
            WHERE id = ?
            """,
            (
                status,
                reject_reason,
                reviewer,
                timestamp,
                None if payload.is_pinned is None else 1 if payload.is_pinned else 0,
                None if payload.is_locked is None else 1 if payload.is_locked else 0,
                timestamp,
                post_id,
            ),
        )
        row = conn.execute(
            """
            SELECT post.*, account.full_name, COUNT(reply.id) AS reply_count
            FROM forum_post AS post
            LEFT JOIN site_account AS account ON account.account = post.author_account
            LEFT JOIN forum_reply AS reply
                ON reply.post_id = post.id
                AND reply.status = ?
                AND COALESCE(reply.deleted_at, '') = ''
            WHERE post.id = ?
            GROUP BY post.id
            """,
            (FORUM_PUBLIC_STATUS, post_id),
        ).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="帖子不存在")
    return forum_post_response(row)


@app.delete("/api/admin/forum/posts/{post_id}")
def delete_forum_post(post_id: str, _: str = Depends(require_admin)) -> dict[str, str]:
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute("SELECT id FROM forum_post WHERE id = ?", (post_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="帖子不存在")
        conn.execute(
            """
            UPDATE forum_post
            SET deleted_at = ?, status = 'hidden', updated_at = ?
            WHERE id = ?
            """,
            (timestamp, timestamp, post_id),
        )
    return {"id": post_id}


@app.put("/api/admin/forum/replies/{reply_id}")
def moderate_forum_reply(
    reply_id: str,
    payload: ForumModerationRequest,
    reviewer: str = Depends(require_admin),
) -> dict[str, str]:
    status = normalize_forum_status(payload.status)
    reject_reason = normalize_forum_content(payload.reject_reason, "处理说明", 500) if payload.reject_reason.strip() else ""
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute("SELECT id, post_id FROM forum_reply WHERE id = ?", (reply_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="回复不存在")
        conn.execute(
            """
            UPDATE forum_reply
            SET status = ?,
                reject_reason = ?,
                reviewed_by = ?,
                reviewed_at = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (status, reject_reason, reviewer, timestamp, timestamp, reply_id),
        )
        conn.execute("UPDATE forum_post SET updated_at = ? WHERE id = ?", (timestamp, existing["post_id"]))
    return {"id": reply_id}


@app.delete("/api/admin/forum/replies/{reply_id}")
def delete_forum_reply(reply_id: str, _: str = Depends(require_admin)) -> dict[str, str]:
    timestamp = now_iso()
    with db_connection() as conn:
        existing = conn.execute("SELECT id, post_id FROM forum_reply WHERE id = ?", (reply_id,)).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="回复不存在")
        conn.execute(
            """
            UPDATE forum_reply
            SET deleted_at = ?, status = 'hidden', updated_at = ?
            WHERE id = ?
            """,
            (timestamp, timestamp, reply_id),
        )
        conn.execute("UPDATE forum_post SET updated_at = ? WHERE id = ?", (timestamp, existing["post_id"]))
    return {"id": reply_id}


@app.get("/api/market/items")
def list_flea_market_items(
    author_account: str = Query(default=""),
    include_delisted: bool = Query(default=False),
    limit: int = Query(default=100, ge=1, le=200),
) -> dict[str, Any]:
    where = ["1 = 1"]
    params: list[Any] = []
    with db_connection() as conn:
        if author_account.strip():
            account = ensure_market_author(conn, author_account)
            where.append("item.author_account = ?")
            params.append(account)
            if not include_delisted:
                where.append("item.status != 'delisted'")
        else:
            where.append("item.status != 'delisted'")
        params.append(limit)
        rows = conn.execute(
            f"""
            SELECT
                item.*,
                account.full_name,
                account.department
            FROM flea_market_item AS item
            LEFT JOIN site_account AS account ON account.account = item.author_account
            WHERE {' AND '.join(where)}
            ORDER BY
                CASE item.status
                    WHEN 'available' THEN 0
                    WHEN 'sold' THEN 1
                    ELSE 2
                END,
                item.created_at DESC
            LIMIT ?
            """,
            params,
        ).fetchall()
    return {"items": [flea_market_item_response(row) for row in rows]}


@app.get("/api/market/items/{item_id}")
def get_flea_market_item(
    item_id: str,
    viewer_account: str = Query(default=""),
) -> dict[str, Any]:
    with db_connection() as conn:
        viewer = ensure_market_author(conn, viewer_account) if viewer_account.strip() else ""
        row = conn.execute(
            """
            SELECT
                item.*,
                account.full_name,
                account.department
            FROM flea_market_item AS item
            LEFT JOIN site_account AS account ON account.account = item.author_account
            WHERE item.id = ?
            """,
            (item_id,),
        ).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="物品不存在")
    if row["status"] == "delisted" and row["author_account"] != viewer:
        raise HTTPException(status_code=404, detail="物品不存在")
    return {"item": flea_market_item_response(row)}


@app.post("/api/market/items")
def create_flea_market_item(payload: FleaMarketItemCreateRequest) -> dict[str, Any]:
    name = normalize_market_name(payload.name)
    image_src = normalize_market_image_src(payload.image_src)
    location = normalize_market_text(payload.location, "存放位置", 120)
    summary = normalize_market_text(payload.summary, "简介", 180)
    detail = normalize_market_text(payload.detail, "详细信息", 2000)
    contact = normalize_market_text(payload.contact, "联系方式", 120)
    tags = normalize_market_tags(payload.tags)
    timestamp = now_iso()
    item_id = uuid.uuid4().hex
    with db_connection() as conn:
        author_account = ensure_market_author(conn, payload.author_account)
        account = conn.execute(
            "SELECT full_name, department FROM site_account WHERE account = ?",
            (author_account,),
        ).fetchone()
        team = account["department"] if account else ""
        conn.execute(
            """
            INSERT INTO flea_market_item (
                id, name, image_src, image_alt, author_account, team, location, status,
                summary, detail, contact, tags, delisted_at, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, 'available', ?, ?, ?, ?, '', ?, ?)
            """,
            (item_id, name, image_src, name, author_account, team, location, summary, detail, contact, tags, timestamp, timestamp),
        )
        row = conn.execute(
            """
            SELECT item.*, account.full_name, account.department
            FROM flea_market_item AS item
            LEFT JOIN site_account AS account ON account.account = item.author_account
            WHERE item.id = ?
            """,
            (item_id,),
        ).fetchone()
    if row is None:
        raise HTTPException(status_code=500, detail="物品发布失败")
    return {"item": flea_market_item_response(row)}


@app.put("/api/market/items/{item_id}")
def update_flea_market_item(item_id: str, payload: FleaMarketItemUpdateRequest) -> dict[str, Any]:
    name = normalize_market_name(payload.name)
    image_src = normalize_market_image_src(payload.image_src)
    location = normalize_market_text(payload.location, "存放位置", 120)
    summary = normalize_market_text(payload.summary, "简介", 180)
    detail = normalize_market_text(payload.detail, "详细信息", 2000)
    contact = normalize_market_text(payload.contact, "联系方式", 120)
    tags = normalize_market_tags(payload.tags)
    timestamp = now_iso()
    with db_connection() as conn:
        author_account = ensure_market_author(conn, payload.author_account)
        existing = conn.execute(
            "SELECT id, author_account FROM flea_market_item WHERE id = ?",
            (item_id,),
        ).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="物品不存在")
        if existing["author_account"] != author_account:
            raise HTTPException(status_code=403, detail="只能编辑自己发布的物品")
        conn.execute(
            """
            UPDATE flea_market_item
            SET name = ?,
                image_src = ?,
                image_alt = ?,
                location = ?,
                summary = ?,
                detail = ?,
                contact = ?,
                tags = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (name, image_src, name, location, summary, detail, contact, tags, timestamp, item_id),
        )
        row = conn.execute(
            """
            SELECT item.*, account.full_name, account.department
            FROM flea_market_item AS item
            LEFT JOIN site_account AS account ON account.account = item.author_account
            WHERE item.id = ?
            """,
            (item_id,),
        ).fetchone()
    return {"item": flea_market_item_response(row)}


@app.put("/api/market/items/{item_id}/status")
def update_flea_market_item_status(item_id: str, payload: FleaMarketStatusRequest) -> dict[str, Any]:
    status = normalize_market_status(payload.status)
    timestamp = now_iso()
    with db_connection() as conn:
        author_account = ensure_market_author(conn, payload.author_account)
        existing = conn.execute(
            "SELECT id, author_account FROM flea_market_item WHERE id = ?",
            (item_id,),
        ).fetchone()
        if existing is None:
            raise HTTPException(status_code=404, detail="物品不存在")
        if existing["author_account"] != author_account:
            raise HTTPException(status_code=403, detail="只能处理自己发布的物品")
        conn.execute(
            """
            UPDATE flea_market_item
            SET status = ?,
                delisted_at = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (status, timestamp if status == "delisted" else "", timestamp, item_id),
        )
        row = conn.execute(
            """
            SELECT item.*, account.full_name, account.department
            FROM flea_market_item AS item
            LEFT JOIN site_account AS account ON account.account = item.author_account
            WHERE item.id = ?
            """,
            (item_id,),
        ).fetchone()
    return {"item": flea_market_item_response(row)}


@app.get("/api/forum/posts")
def list_forum_posts() -> dict[str, Any]:
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                post.*,
                account.full_name,
                COUNT(reply.id) AS reply_count
            FROM forum_post AS post
            LEFT JOIN site_account AS account ON account.account = post.author_account
            LEFT JOIN forum_reply AS reply
                ON reply.post_id = post.id
                AND reply.status = ?
                AND COALESCE(reply.deleted_at, '') = ''
            WHERE post.status = ?
                AND COALESCE(post.deleted_at, '') = ''
            GROUP BY post.id
            ORDER BY post.is_pinned DESC, post.created_at DESC
            LIMIT 100
            """,
            (FORUM_PUBLIC_STATUS, FORUM_PUBLIC_STATUS),
        ).fetchall()
    return {"posts": [forum_post_response(row) for row in rows]}


@app.get("/api/forum/inbox")
def list_forum_inbox(author_account: str = Query(...)) -> dict[str, Any]:
    with db_connection() as conn:
        account = ensure_forum_author(conn, author_account)
        posts = conn.execute(
            """
            SELECT
                post.*,
                account.full_name,
                COUNT(reply.id) AS reply_count
            FROM forum_post AS post
            LEFT JOIN site_account AS account ON account.account = post.author_account
            LEFT JOIN forum_reply AS reply
                ON reply.post_id = post.id
                AND reply.status = ?
                AND COALESCE(reply.deleted_at, '') = ''
            WHERE post.author_account = ?
                AND post.status IN ('pending', 'rejected')
                AND COALESCE(post.deleted_at, '') = ''
            GROUP BY post.id
            ORDER BY post.updated_at DESC
            LIMIT 100
            """,
            (FORUM_PUBLIC_STATUS, account),
        ).fetchall()
        replies = conn.execute(
            """
            SELECT
                reply.*,
                account.full_name,
                post.title AS post_title,
                post.status AS post_status
            FROM forum_reply AS reply
            LEFT JOIN site_account AS account ON account.account = reply.author_account
            LEFT JOIN forum_post AS post ON post.id = reply.post_id
            WHERE reply.author_account = ?
                AND reply.status IN ('pending', 'rejected')
                AND COALESCE(reply.deleted_at, '') = ''
            ORDER BY reply.updated_at DESC
            LIMIT 100
            """,
            (account,),
        ).fetchall()
    reply_payload = []
    for row in replies:
        item = forum_reply_response(row)
        item["post_title"] = row["post_title"] or ""
        item["post_status"] = row["post_status"] or ""
        reply_payload.append(item)
    return {
        "posts": [forum_post_response(row) for row in posts],
        "replies": reply_payload,
    }


@app.get("/api/forum/my-posts/{post_id}")
def get_my_forum_post(post_id: str, author_account: str = Query(...)) -> dict[str, Any]:
    with db_connection() as conn:
        account = ensure_forum_author(conn, author_account)
        post = conn.execute(
            """
            SELECT
                post.*,
                account.full_name,
                COUNT(reply.id) AS reply_count
            FROM forum_post AS post
            LEFT JOIN site_account AS account ON account.account = post.author_account
            LEFT JOIN forum_reply AS reply
                ON reply.post_id = post.id
                AND reply.status = ?
                AND COALESCE(reply.deleted_at, '') = ''
            WHERE post.id = ?
                AND post.author_account = ?
                AND COALESCE(post.deleted_at, '') = ''
            GROUP BY post.id
            """,
            (FORUM_PUBLIC_STATUS, post_id, account),
        ).fetchone()
        if post is None:
            raise HTTPException(status_code=404, detail="post not found")
    return {
        "post": forum_post_response(post),
        "replies": [],
    }


@app.post("/api/forum/posts")
def create_forum_post(payload: ForumPostCreateRequest) -> dict[str, Any]:
    title = normalize_forum_title(payload.title)
    content = normalize_forum_content(payload.content)
    timestamp = now_iso()
    post_id = uuid.uuid4().hex
    with db_connection() as conn:
        author_account = ensure_forum_author(conn, payload.author_account)
        conn.execute(
            """
            INSERT INTO forum_post (
                id, title, content, author_account, status, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, 'pending', ?, ?)
            """,
            (post_id, title, content, author_account, timestamp, timestamp),
        )
        row = conn.execute(
            """
            SELECT post.*, account.full_name, 0 AS reply_count
            FROM forum_post AS post
            LEFT JOIN site_account AS account ON account.account = post.author_account
            WHERE post.id = ?
            """,
            (post_id,),
        ).fetchone()
    if row is None:
        raise HTTPException(status_code=500, detail="帖子创建失败")
    return forum_post_response(row)


@app.put("/api/forum/posts/{post_id}")
def update_forum_post(post_id: str, payload: ForumPostUpdateRequest) -> dict[str, Any]:
    title = normalize_forum_title(payload.title)
    content = normalize_forum_content(payload.content)
    timestamp = now_iso()
    with db_connection() as conn:
        author_account = ensure_forum_author(conn, payload.author_account)
        existing = conn.execute(
            """
            SELECT id, author_account, deleted_at
            FROM forum_post
            WHERE id = ?
            """,
            (post_id,),
        ).fetchone()
        if existing is None or existing["deleted_at"]:
            raise HTTPException(status_code=404, detail="帖子不存在")
        if existing["author_account"] != author_account:
            raise HTTPException(status_code=403, detail="只能编辑自己发布的帖子")
        conn.execute(
            """
            UPDATE forum_post
            SET title = ?,
                content = ?,
                status = 'pending',
                reject_reason = '',
                reviewed_by = '',
                reviewed_at = '',
                updated_at = ?
            WHERE id = ?
            """,
            (title, content, timestamp, post_id),
        )
        row = conn.execute(
            """
            SELECT post.*, account.full_name, COUNT(reply.id) AS reply_count
            FROM forum_post AS post
            LEFT JOIN site_account AS account ON account.account = post.author_account
            LEFT JOIN forum_reply AS reply
                ON reply.post_id = post.id
                AND reply.status = ?
                AND COALESCE(reply.deleted_at, '') = ''
            WHERE post.id = ?
            GROUP BY post.id
            """,
            (FORUM_PUBLIC_STATUS, post_id),
        ).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="帖子不存在")
    return forum_post_response(row)


@app.get("/api/forum/posts/{post_id}")
def get_forum_post(post_id: str) -> dict[str, Any]:
    with db_connection() as conn:
        post = conn.execute(
            """
            SELECT
                post.*,
                account.full_name,
                COUNT(reply.id) AS reply_count
            FROM forum_post AS post
            LEFT JOIN site_account AS account ON account.account = post.author_account
            LEFT JOIN forum_reply AS reply
                ON reply.post_id = post.id
                AND reply.status = ?
                AND COALESCE(reply.deleted_at, '') = ''
            WHERE post.id = ?
                AND post.status = ?
                AND COALESCE(post.deleted_at, '') = ''
            GROUP BY post.id
            """,
            (FORUM_PUBLIC_STATUS, post_id, FORUM_PUBLIC_STATUS),
        ).fetchone()
        if post is None:
            raise HTTPException(status_code=404, detail="帖子不存在")
        replies = conn.execute(
            """
            SELECT reply.*, account.full_name
            FROM forum_reply AS reply
            LEFT JOIN site_account AS account ON account.account = reply.author_account
            WHERE reply.post_id = ?
                AND reply.status = ?
                AND COALESCE(reply.deleted_at, '') = ''
            ORDER BY reply.created_at ASC
            """,
            (post_id, FORUM_PUBLIC_STATUS),
        ).fetchall()
    return {
        "post": forum_post_response(post),
        "replies": [forum_reply_response(row) for row in replies],
    }


@app.post("/api/forum/posts/{post_id}/replies")
def create_forum_reply(post_id: str, payload: ForumReplyCreateRequest) -> dict[str, str]:
    content = normalize_forum_content(payload.content, "回复", 2000)
    timestamp = now_iso()
    reply_id = uuid.uuid4().hex
    with db_connection() as conn:
        author_account = ensure_forum_author(conn, payload.author_account)
        post = conn.execute(
            """
            SELECT id, is_locked
            FROM forum_post
            WHERE id = ?
                AND status = ?
                AND COALESCE(deleted_at, '') = ''
            """,
            (post_id, FORUM_PUBLIC_STATUS),
        ).fetchone()
        if post is None:
            raise HTTPException(status_code=404, detail="帖子不存在")
        if post["is_locked"]:
            raise HTTPException(status_code=400, detail="帖子已锁定，暂时不能回复")
        conn.execute(
            """
            INSERT INTO forum_reply (id, post_id, content, author_account, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, 'pending', ?, ?)
            """,
            (reply_id, post_id, content, author_account, timestamp, timestamp),
        )
        conn.execute("UPDATE forum_post SET updated_at = ? WHERE id = ?", (timestamp, post_id))
    return {"id": reply_id, "post_id": post_id}


@app.put("/api/forum/replies/{reply_id}")
def update_forum_reply(reply_id: str, payload: ForumReplyUpdateRequest) -> dict[str, Any]:
    content = normalize_forum_content(payload.content, "reply", 2000)
    timestamp = now_iso()
    with db_connection() as conn:
        author_account = ensure_forum_author(conn, payload.author_account)
        existing = conn.execute(
            """
            SELECT id, post_id, author_account, deleted_at
            FROM forum_reply
            WHERE id = ?
            """,
            (reply_id,),
        ).fetchone()
        if existing is None or existing["deleted_at"]:
            raise HTTPException(status_code=404, detail="reply not found")
        if existing["author_account"] != author_account:
            raise HTTPException(status_code=403, detail="cannot edit this reply")
        conn.execute(
            """
            UPDATE forum_reply
            SET content = ?,
                status = 'pending',
                reject_reason = '',
                reviewed_by = '',
                reviewed_at = '',
                updated_at = ?
            WHERE id = ?
            """,
            (content, timestamp, reply_id),
        )
        conn.execute("UPDATE forum_post SET updated_at = ? WHERE id = ?", (timestamp, existing["post_id"]))
        row = conn.execute(
            """
            SELECT reply.*, account.full_name
            FROM forum_reply AS reply
            LEFT JOIN site_account AS account ON account.account = reply.author_account
            WHERE reply.id = ?
            """,
            (reply_id,),
        ).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="reply not found")
    return forum_reply_response(row)


@app.post("/api/auth/login")
def login(payload: LoginRequest) -> dict[str, str]:
    if payload.role == "admin" and payload.password == ADMIN_PASSWORD:
        return {"token": make_token("admin"), "role": "admin"}
    raise HTTPException(status_code=401, detail="密码错误")


@app.get("/api/season-plan")
def get_season_plan(season_year: int = 2026, month: int = 6) -> dict[str, Any]:
    return {
        "season_year": season_year,
        "month": month,
        "plans": list_season_plan(season_year, month),
    }


@app.put("/api/season-plan")
def update_season_plan(
    payload: SeasonPlanRequest,
) -> dict[str, Any]:
    require_site_plan_editor(payload)
    return {
        "season_year": payload.season_year,
        "month": payload.month,
        "plans": save_season_plan(payload.season_year, payload.month, payload.plans),
    }


@app.post("/api/invoices/upload")
async def upload_invoice_form(
    team_name: str = Form(...),
    submitter_name: str = Form(...),
    remark: str = Form(""),
    form_file: UploadFile = File(...),
) -> dict[str, str]:
    if not team_name.strip() or not submitter_name.strip():
        raise HTTPException(status_code=400, detail="兵种名称、提交人姓名不能为空")
    if not form_file.filename or Path(form_file.filename).suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="只支持上传 .xlsx 表格")
    content = await form_file.read()
    if len(content) > MAX_CONTENT_LENGTH:
        raise HTTPException(status_code=413, detail="文件超过 60MB 上限")

    submitter_id = "invoice"
    batch_id = build_batch_id(submitter_id)
    batch_dir = UNREGISTERED_DIR / batch_id
    batch_dir.mkdir(parents=True, exist_ok=True)
    target_path = batch_dir / "form.xlsx"
    target_path.write_bytes(content)
    submitted_at = now_iso()
    write_meta(
        batch_dir,
        {
            "batch_id": batch_id,
            "team_name": team_name,
            "submitter_name": submitter_name,
            "submitter_id": submitter_id,
            "submitted_at": submitted_at,
            "form_file": "form.xlsx",
            "remark": remark,
            "folder_stage": "unregistered",
        },
    )
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO upload_batch (
                id, team_name, submitter_name, submitter_id, submitted_at,
                form_file_path, remark, status, folder_stage, review_note, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (batch_id, team_name, submitter_name, submitter_id, submitted_at, str(target_path), remark, "unregistered", "unregistered", "", submitted_at, submitted_at),
        )
    log_process("upload", "info", "上传成功，等待本地审核脚本处理", batch_id)
    return {"batch_id": batch_id, "message": "上传成功"}


@app.get("/api/invoices/template")
def download_template() -> StreamingResponse:
    stream = build_template_workbook()
    filename = "发票表格模板.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@app.get("/api/invoices/dashboard")
def get_dashboard(_: str = Depends(require_admin)) -> dict[str, Any]:
    return dashboard_data()


@app.post("/api/invoices/review/run")
def run_review(_: str = Depends(require_admin)) -> dict[str, int]:
    return {"processed": run_agent_once()}


@app.get("/api/invoices/batches/{batch_id}")
def get_batch_detail(batch_id: str, _: str = Depends(require_admin)) -> dict[str, Any]:
    with db_connection() as conn:
        batch = conn.execute("SELECT * FROM upload_batch WHERE id = ?", (batch_id,)).fetchone()
        if batch is None:
            raise HTTPException(status_code=404, detail="批次不存在")
        rows = conn.execute("SELECT * FROM staged_purchase_record WHERE batch_id = ? ORDER BY row_no ASC", (batch_id,)).fetchall()
    return {"batch": row_to_dict(batch), "rows": [row_to_dict(row) for row in rows]}


@app.post("/api/invoices/batches/{batch_id}/confirm-selected")
def confirm_selected(batch_id: str, payload: RowIdsRequest, _: str = Depends(require_admin)) -> dict[str, int]:
    return {"count": confirm_staged_rows(batch_id, payload.row_ids)}


@app.post("/api/invoices/batches/{batch_id}/confirm-all")
def confirm_all(batch_id: str, _: str = Depends(require_admin)) -> dict[str, int]:
    with db_connection() as conn:
        rows = conn.execute(
            "SELECT id FROM staged_purchase_record WHERE batch_id = ? AND review_status = 'pending_review'",
            (batch_id,),
        ).fetchall()
    return {"count": confirm_staged_rows(batch_id, [row["id"] for row in rows])}


@app.post("/api/invoices/batches/{batch_id}/reject-selected")
def reject_selected(batch_id: str, payload: RowIdsRequest, _: str = Depends(require_admin)) -> dict[str, int]:
    note = payload.note.strip() or "人工复核未通过"
    return {"count": reject_staged_rows(batch_id, payload.row_ids, note)}


@app.post("/api/invoices/batches/{batch_id}/complete")
def complete_batch(batch_id: str, _: str = Depends(require_admin)) -> dict[str, Any]:
    success, message = finalize_batch_review(batch_id)
    if not success:
        raise HTTPException(status_code=400, detail=message)
    return {"message": message}


@app.post("/api/invoices/reimburse")
def reimburse(payload: ReimburseRequest, _: str = Depends(require_admin)) -> dict[str, str]:
    success, message = export_reimbursement(payload.record_ids)
    if not success:
        raise HTTPException(status_code=400, detail=message)
    return {"reimbursement_id": message, "message": "出库完成"}


@app.get("/api/invoices/forms/{batch_id}")
def download_form(batch_id: str, _: str = Depends(require_admin)) -> FileResponse:
    with db_connection() as conn:
        batch = conn.execute("SELECT * FROM upload_batch WHERE id = ?", (batch_id,)).fetchone()
    if batch is None:
        raise HTTPException(status_code=404, detail="批次不存在")
    path = Path(batch["form_file_path"])
    if not path.exists():
        path = batch_form_path(batch_id, batch["folder_stage"])
    if not path.exists():
        raise HTTPException(status_code=404, detail="表格文件不存在")
    return FileResponse(path, filename=f"{batch_id}.xlsx")


@app.get("/api/invoices/reimbursements/{reimbursement_id}")
def download_reimbursement(reimbursement_id: str, _: str = Depends(require_admin)) -> FileResponse:
    with db_connection() as conn:
        batch = conn.execute("SELECT * FROM reimbursement_batch WHERE id = ?", (reimbursement_id,)).fetchone()
    if batch is None:
        raise HTTPException(status_code=404, detail="出库批次不存在")
    path = Path(batch["export_file_path"])
    if not path.exists():
        raise HTTPException(status_code=404, detail="导出文件不存在")
    return FileResponse(path, filename=f"{reimbursement_id}.xlsx")
